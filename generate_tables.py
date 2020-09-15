from copy import deepcopy
from copy import copy
import logging
import os.path
from io import BytesIO
from io import StringIO
from datetime import datetime  #to get current time
import platform

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
import pandas as pd
from tabulate import tabulate

import cpuc.params as params
from cpuc.sharefileapi import ShareFileSession
from cpuc.sharefileapi import SHAREFILE_OPTIONS
from cpuc.mylogging import create_logfile
from cpuc.tables import OldTable
from cpuc.tables import Table
from read_table_spec import gettablespecs
from cpuc.utility import formatws
from cpuc.utility import excel_to_df
from cpuc.wincom import openWord
from cpuc.googleaccess import download_drive_file
#hopefully this is temporary
import create_ciac_2018_report_data_files as CIAC2018

DATA_FILE = os.path.join(
    params.SAMPLED_SITE_REVIEW_PATH,
    'evaldata.xlsx',
)

TEMPLATE_FILE = os.path.join(
    params.D0_DELIVERABLES_PATH,
    'Report',
    'Drafts',
    '2019-5-15',
    'D0 Report Figures and Tables - Design.xlsx',
)

TEMPLATE_FILE_OLD = os.path.join(
    params.D0_DELIVERABLES_PATH,
    'Report',
    'Drafts',
    'testing',
    'D0 Report Figures and Tables - Design.xlsx',
)

TEMPLATE_FILE_D1_TEST = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\01 - 2018-19 Ex Post Workplans\IALC\Workplan\Drafts\2019-7-9\D1 IALC Workplan Figures and Tables - Design.xlsx'

WORD_TEMPLATE = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\00 - 2017 Evaluation\Report\Drafts\2019-5-15\CPUC_Table_Template.docx'
EXPORT_FILE_D1_TEST = r''
EXPORT_FILE = os.path.join(
    params.D0_DELIVERABLES_PATH,
    'Report',
    'Drafts',
    'testing',
    'D0 Report Figures and Tables - Report.xlsx',
)

WORD_TABLE_TEMPLATE = os.path.join(
    params.DELIVERABLES_PATH,
    '__ - Cross Cutting',
    'Writing Tools',
    'Template',
    'Excel to Word',
    'CPUC_Table_Template.docx',
)

STEPS_DBR_COUNTS = os.path.join(
    params.SAMPLED_SITE_REVIEW_PATH,
    'steps_dbr_CountsbyPA.csv',
)
STEPS_DBR_SUMS = os.path.join(
    params.SAMPLED_SITE_REVIEW_PATH,
    'steps_dbr_SumsbyPA.csv',
)
STEPS_DS_COUNTS = os.path.join(
    params.SAMPLED_SITE_REVIEW_PATH,
    'steps_DS_CountsbyPA.csv',
)
STEPS_DS_SUMS = os.path.join(
    params.SAMPLED_SITE_REVIEW_PATH,
    'steps_DS_SumsbyPA.csv',
)

PA_MAP = {
    'PGE': 'PG&E',
    'SCE': 'SCE',
    'SDGE': 'SDG&E',
    'SCG': 'SCG',
}


def build_tablespecs():
    wb = load_workbook(TEMPLATE_FILE, data_only=False)
    for ws in wb.worksheets:
        if ws.title in ['Captions', 'Tools']:
            continue

        print(ws.merged_cells.ranges)
        

        print('*** Worksheet {}'.format(ws.title))
        print(ws['C1'].value)

        right = 0
        for cell in ws['1']:
            if cell.value == 'rowspec':
                right = cell.column
        bottom = 0
        header_row = 0
        for cell in ws['A']:
            if cell.value == 'headers':
                header_row = cell.row
            if cell.value == 'colspec':
                bottom = cell.row

        print('right: {}, bottom: {}'.format(right, bottom))
        if not (right and bottom and header_row):
            print('No table found in worksheet {}'.format(ws.title))
            print()
            continue

        tablespec = {}
        sections = []
        colspec = {}
        row_index = 2
        col_index = 2
        first_header_col = 0

        tablespec['title'] = ws['C1'].value
        tablespec['name'] = ws.title

        colspec['field'] = ws.cell(row=bottom, column=1).value
        colspec['columns'] = []

        for row in ws.iter_rows(min_row=2, min_col=2, max_row=bottom, max_col=right):
            cells = []
            for cell in row:
                row_index = cell.row
                col_index = cell.column

                if row_index == header_row:
                    if (not first_header_col) and cell.value:
                        first_header_col = cell.column
                    if cell.value:
                        colspec['columns'].append({'caption': cell.value})

                value = None
                if cell.alignment.horizontal:
                    value = '{}: {}'.format(cell.alignment.horizontal, cell.value)
                else:
                    value = cell.value
                cells.append(value)
            print(tuple(cells))
        print()

        tablespec['colspec'] = colspec
        tablespec['sections'] = sections
        return tablespec


projects_and_claims_spec = {
    'name': '1',
    'title': 'Database and Sampled Projects and Claims by PA',
    'colspec': {
        'field': 'PA',
        'columns': [
            {
                'caption': 'PG&E',
                'value': 'PGE',
            },
            {
                'caption': 'SCE',
                'value':  'SCE',
            },
            {
                'caption': 'SDG&E',
                'value': 'SDGE',
            },
            {
                'caption': 'SCG',
                'value': 'SCG',
            },
            {
                'caption': 'Total',
                'value': 'total',
            },
        ],
    },
    'sections': [
        {
            'header': 'Projects',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['project_count'],
                        },
                    ],
                },
                {
                    'index': 'Sampled',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['project_count'],
                        },
                    ],
                },
                {
                    'index': '% in Sample',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['project_count'],
                        },
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['project_count'],
                        },
                    ],
                    'reducer': OldTable.percent_of_first,
                },
            ],
        },
        {
            'header': 'Claims',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['count'],
                        },
                    ],
                },
                {
                    'index': 'Sampled',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['count'],
                        },
                    ],
                },
                {
                    'index': '% in Sample',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['count'],
                        },
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['count'],
                        },
                    ],
                    'reducer': OldTable.percent_of_first,
                },
            ],
        },
        {
            'header': 'Claims per Project',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['count'],
                        },
                        {
                            'field': 'SampledProject',
                            'filters': ['all'],
                            'transforms': ['project_count'],
                        },
                    ],
                    'reducer': OldTable.ratio,
                },
                {
                    'index': 'Sampled',
                    'fields': [
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['count'],
                        },
                        {
                            'field': 'SampledProject',
                            'filters': ['is_sampled'],
                            'transforms': ['project_count'],
                        },
                    ],
                    'reducer': OldTable.ratio,
                },
            ],
        },
    ],
}

claimed_vs_sampled_savings_spec = {
    'name': '2',
    'title': 'Claimed Lifetime Net Savings and Percent Sampled by PA',
    'colspec': {
        'field': 'PA',
        'columns': [
            {
                'caption': 'PG&E',
                'value': 'PGE',
            },
            {
                'caption': 'SCE',
                'value':  'SCE',
            },
            {
                'caption': 'SDG&E',
                'value': 'SDGE',
            },
            {
                'caption': 'SCG',
                'value': 'SCG',
            },
            {
                'caption': 'Total',
                'value': 'total',
            },
        ],
    },
    'sections': [
        {
            'header': 'MWh',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data', 'kWh_to_MWh'],
                        },
                    ],
                },
                {
                    'index': 'Project Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data', 'kWh_to_MWh'],
                        },
                    ],
                },
                {
                    'index': '% in Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data', 'kWh_to_MWh'],
                        },
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data', 'kWh_to_MWh'],
                        },
                    ],
                    'reducer': OldTable.percent_of_first,
                },
            ]
        },
        {
            'header': 'MW',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data', 'kW_to_MW'],
                        },
                    ],
                },
                {
                    'index': 'Project Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data', 'kW_to_MW'],
                        },
                    ],
                },
                {
                    'index': '% in Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data', 'kW_to_MW'],
                        },
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data', 'kW_to_MW'],
                        },
                    ],
                    'reducer': OldTable.percent_of_first,
                },
            ]
        },
        {
            'header': 'Therms',
            'rows': [
                {
                    'index': 'Claims Database',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data'],
                        },
                    ],
                },
                {
                    'index': 'Project Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data'],
                        },
                    ],
                },
                {
                    'index': '% in Sample',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'filters': ['all'],
                            'transforms': ['take_field', 'sum_data'],
                        },
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'filters': ['is_sampled'],
                            'transforms': ['take_field', 'sum_data'],
                        },
                    ],
                    'reducer': OldTable.percent_of_first,
                },
            ]
        },
    ]
}

change_in_savings_spec = {
    'name': '4',
    'title': 'Change in Savings Due to Database (DB) and Sample Review by PA',
    'colspec': {
        'field': 'PA',
        'columns': [
            {
                'caption': 'PG&E',
                'value': 'PGE',
            },
            {
                'caption': 'SCE',
                'value':  'SCE',
            },
            {
                'caption': 'SDG&E',
                'value': 'SDGE',
            },
            {
                'caption': 'SCG',
                'value': 'SCG',
            },
            {
                'caption': 'Total',
                'value': 'total',
            },
        ],
    },
    'sections': [
        {
            'header': 'MWh',
            'rows': [
                {
                    'index': 'Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                    ],
                },
                {
                    'index': 'DB Review',
                    'fields': [
                        {
                            'field': 'cdrExPostLifecycleNetkWh',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                        {
                            'field': 'cdrExPostLifecycleNetkWh',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
                {
                    'index': 'DB and Sample Review',
                    'fields': [
                        {
                            'field': 'EvalExPostLifecycleNetkWh_atr',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkWh',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                        {
                            'field': 'EvalExPostLifecycleNetkWh_atr',
                            'transforms': ['sum_data', 'kWh_to_MWh'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
            ]
        },
        {
            'header': 'MW',
            'rows': [
                {
                    'index': 'Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                    ],
                },
                {
                    'index': 'DB Review',
                    'fields': [
                        {
                            'field': 'cdrExPostLifecycleNetkW',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                        {
                            'field': 'cdrExPostLifecycleNetkW',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
                {
                    'index': 'DB and Sample Review',
                    'fields': [
                        {
                            'field': 'EvalExPostLifecycleNetkW_atr',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetkW',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                        {
                            'field': 'EvalExPostLifecycleNetkW_atr',
                            'transforms': ['sum_data', 'kW_to_MW'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
            ]
        },
        {
            'header': 'Therms',
            'rows': [
                {
                    'index': 'Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'transforms': ['sum_data'],
                        },
                    ],
                },
                {
                    'index': 'DB Review',
                    'fields': [
                        {
                            'field': 'cdrExPostLifecycleNetTherm',
                            'transforms': ['sum_data'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'transforms': ['sum_data'],
                        },
                        {
                            'field': 'cdrExPostLifecycleNetTherm',
                            'transforms': ['sum_data'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
                {
                    'index': 'DB and Sample Review',
                    'fields': [
                        {
                            'field': 'EvalExPostLifecycleNetTherm_atr',
                            'transforms': ['sum_data'],
                        },
                    ],
                },
                {
                    'index': '% change from Claim',
                    'fields': [
                        {
                            'field': 'ExAnteLifecycleNetTherm',
                            'transforms': ['sum_data'],
                        },
                        {
                            'field': 'EvalExPostLifecycleNetTherm_atr',
                            'transforms': ['sum_data'],
                        },
                    ],
                    'reducer': OldTable.percent_change,
                },
            ]
        },
    ]
}

table5 = [
    {
        'index': None,
        'passthrough': True,
        'columns': [
            'MWh', 'MW', 'Therms', 'MWh', 'MW', 'Therms',
        ],
    },
    {
        'index': 'Reported Claims',
        'columns': [
            'total_claims',
            'total_claims',
            'total_claims',
            'ExAnteLifecycleNetkWh',
            'ExAnteLifecycleNetkW',
            'ExAnteLifecycleNetTherm',
        ],
    },
    {
        'index': 'Database Review',
        'columns': [None, None, None, None, None, None,],
    },
    {
        'index': 'Installation Date',
        'columns': [
            'cdrdateineligibleflagkwh',
            'cdrdateineligibleflagkw',
            'cdrdateineligibleflagthm',
            'cdrdatekwh',
            'cdrdatekw',
            'cdrdatethm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'cdrdatekwh',
            'cdrdatekw',
            'cdrdatethm',
        ],
    },
    {
        'index': 'Net-to-Gross Ratio',
        'columns': [
            'cdrdatentgineligibleflagkwh',
            'cdrdatentgineligibleflagkw',
            'cdrdatentgineligibleflagthm',
            'cdrntgeligkwh',
            'cdrntgeligkw',
            'cdrntgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'cdrntgeligkwh',
            'cdrntgeligkw',
            'cdrntgeligthm',
        ],
    },
    {
        'index': 'EUL or RUL',
        'columns': [
            'cdrdatentgulineligibleflagkwh',
            'cdrdatentgulineligibleflagkw',
            'cdrdatentgulineligibleflagthm',
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': 'Gross Realization Rate',
        'columns': [
            'cdrdatentgulineligibleflagkwh',
            'cdrdatentgulineligibleflagkw',
            'cdrdatentgulineligibleflagthm',
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': 'Installation Rate',
        'columns': [
            'cdrdatentgulineligibleflagkwh',
            'cdrdatentgulineligibleflagkw',
            'cdrdatentgulineligibleflagthm',
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'cdrulntgeligkwh',
            'cdrulntgeligkw',
            'cdrulntgeligthm',
        ],
    },
    {
        'index': 'Database and Sample Review',
        'columns': [None, None, None, None, None, None,],
    },
    {
        'index': 'Eligibility',
        'columns': [
            'atr_ineligibleflagkwh',
            'atr_ineligibleflagkw',
            'atr_ineligibleflagthm',
            'atr_eligkwh',
            'atr_eligkw',
            'atr_eligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'atr_eligkwh',
            'atr_eligkw',
            'atr_eligthm',
        ],
    },
    {
        'index': 'Gross Savings',
        'columns': [
            'atr_svgschangeflagkwh',
            'atr_svgschangeflagkw',
            'atr_svgschangeflagthm',
            'atr_svgsEligkwh',
            'atr_svgsEligkw',
            'atr_svgsEligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'atr_svgsEligkwh',
            'atr_svgsEligkw',
            'atr_svgsEligthm',
        ],
    },
    {
        'index': 'Net-to-Gross Ratio',
        'columns': [
            'atr_NTGchangeflagkwh',
            'atr_NTGchangeflagkw',
            'atr_NTGchangeflagthm',
            'atr_NTGsvgeligkwh',
            'atr_NTGsvgeligkw',
            'atr_NTGsvgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'atr_NTGsvgeligkwh',
            'atr_NTGsvgeligkw',
            'atr_NTGsvgeligthm',
        ],
    },
    {
        'index': 'EUL or RUL',
        'columns': [
            'atr_ULchangeflagkwh',
            'atr_ULchangeflagkw',
            'atr_ULchangeflagthm',
            'atr_NTGULsvgeligkwh',
            'atr_NTGULsvgeligkw',
            'atr_NTGULsvgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'atr_NTGULsvgeligkwh',
            'atr_NTGULsvgeligkw',
            'atr_NTGULsvgeligthm',
        ],
    },
    {
        'index': 'All Changes',
        'columns': [None, None, None, None, None, None,],
    },
    {
        'index': 'All Changes',
        'columns': [
            'atr_ULchangeflagkwh',
            'atr_ULchangeflagkw',
            'atr_ULchangeflagthm',
            'atr_NTGULsvgeligkwh',
            'atr_NTGULsvgeligkw',
            'atr_NTGULsvgeligthm',
        ],
    },
    {
        'index': '% change from claim',
        'columns': [
            None,
            None,
            None,
            'atr_NTGULsvgeligkwh',
            'atr_NTGULsvgeligkw',
            'atr_NTGULsvgeligthm',
        ],
    },
]

table6 = [
    {
        'index': None,
        'passthrough': True,
        'value': 'Claims',
    },
    {
        'index': 'Reported Claims',
        'value': "len(data[data['SampledProject'] == 1][data.PA == '{pa}'])",
    },
    {
        'index': 'Savings Set to Zero',
        'value': None,
    },
    {
        'index': 'Install date before 2017 without M&V justification',
        'value': "len(data[data.PA == '{pa}'][data['RvwInstallDate'] == 'No'])",
    },
    {
        'index': 'Application signed after install date',
        'value': "len(data[data.PA == '{pa}'][data['RvwAppVsInstallDate'] == 'No'])",
    },
    {
        'index': 'No evidence that required permits were obtained',
        'value': "len(data[data.PA == '{pa}'][data['RvwPermit'] == 'No'])",
    },
    {
        'index': 'Did not pass fuel switching 3-prong test',
        'value': "len(data[data.PA == '{pa}'][data['RvwFuelSwitchTest'] == 'No'])",
    },
    {
        'index': 'Savings not adjusted for on-site generation',
        'value': "len(data[data.PA == '{pa}'][data['RvwCogenImpact'] == 'No'])",
    },
    {
        'index': 'Measure efficiency did not exceed that of existing equipment',
        'value': "len(data[data.PA == '{pa}'][data['RvwEffIncrease'] == 'No'])",
    },
    {
        'index': 'Other Savings Adjustments',
        'value': None,
    },
    {
        'index': 'No evidence of program influence',
        'value': "len(data[data.PA == '{pa}'][data['ProgInfluenceFlag'] == 1])",
    },
    {
        'index': 'Changed measure application type',
        'value': None,
    },
    {
        'index': 'No justification provided for early replacement',
        'value': "len(data[data.PA == '{pa}'][data['NoERJust'] == 1][data['MeasAppType'] == 'ER'][data['EvalBaselineType'] != 'ER'])",
    },
    {
        'index': 'RUL less than one year',
        'value': "len(data[data.PA == '{pa}'][data['RULUnder1'] == 1])",
    },
    {
        'index': 'Did not exceed all applicable codes and regulations',
        'value': "len(data[data.PA == '{pa}'][data['RvwCodeRegs'] == 'No'])",
    },
    {
        'index': 'Did not exceed Industry Standard Practices',
        'value': "len(data[data.PA == '{pa}'][data['RvwISPMet'] == 'No'])",
    },
]

def style_table1(raw):
    rows = deepcopy(raw)

    # Add heading format to first column.
    for row in rows:
        if row[0]:
            if row[1:] == [None, None, None, None, None]:
                row[0] = 'style: tbl head|{}'.format(row[0])
                # TODO: exclude % rows, or fill them in so they aren't caught in
                # this first if clause
            elif row[0].startswith('%'):
                row[0] = 'style: tbl head r|{}'.format(row[0])
            else:
                row[0] = 'style: tbl head|\t{}'.format(row[0])

    # Insert 5-column span label
    rows.insert(0, [None, 'colspan: 5, style: tbl head c|Claimed Lifetime Net Savings'])

    # Add heading style to unit headers
    for index, cell in enumerate(rows[1]):
        if cell:
            rows[1][index] = 'style: tbl head r|{}'.format(cell)
    return rows

def get_table5_data():
    dbr_counts = pd.read_csv(STEPS_DBR_COUNTS)
    dbr_sums = pd.read_csv(STEPS_DBR_SUMS)
    ds_counts = pd.read_csv(STEPS_DS_COUNTS)
    ds_sums = pd.read_csv(STEPS_DS_SUMS)

    dbr_data = pd.merge(dbr_counts, dbr_sums)
    ds_data = pd.merge(ds_counts, ds_sums)

    return pd.merge(dbr_data, ds_data)

def build_table5(data, all_claims):
    rows = []
    for pa in ['PGE', 'SCE', 'SDGE', 'SCG']:
        rows += build_table5_pa(data, all_claims, pa)
    return rows

def build_table5_pa(data, all_claims, pa):
    logging.info('Building table 5: %s' % pa)
    rows = []
    rows.append([pa] + [None] * 6)
    pa_data = data[data['PA'] == pa]

    reported = [None] * 6
    for row in table5:
        cols = []
        for col_index, col in enumerate(row['columns']):
            if row.get('passthrough'):
                cols.append(col)
            elif col == 'total_claims':
                cols.append(len(all_claims[all_claims['PA'] == pa]))
            else:
                if col:
                    value = float(data.loc[data.PA.values == pa, col])
                    if 'kw' in col.lower():
                        value /= 1000
                    if row['index'].startswith('%') and reported[col_index]:
                        value = (value - reported[col_index]) / reported[col_index] * 100
                    cols.append(value)

                    # Store reported claims for % calculations
                    if row['index'] == 'Reported Claims':
                        reported[col_index] = value
                else:
                    cols.append(None)
        rows.append(
            [row['index']] + cols
        )
    rows.append([None] * 7)
    return rows

def style_table5(raw):
    rows = deepcopy(raw)
    pa = rows[0][0]
    logging.info('Styling table 5: %s' % pa)

    # Remove first and last rows; they're intended to separate PAs in a list
    # of all 4 PA tables.
    del rows[0]
    del rows[-1]

    for row in rows:
        # Strip out first three data columns (eligibility flags, no
        # longer used).
        del row[1:4]

        # Add heading format to first column.
        # Add bold format to appropriate data cells.
        if row[0]:
            if row[0].startswith('%'):
                row[0] = [{
                    'style': 'tbl strong r',
                    'background': '#eaf1dd',
                }, row[0]]
                for index, col in enumerate(row[1:]):
                    row[index + 1] = [{
                        'style': 'tbl strong r',
                        'percent': True,
                    }, row[index + 1]]
            elif row[1:] == [None, None, None]:
                row[0] = [{
                    'style': 'tbl head',
                    'background': '#eaf1dd',
                }, row[0]]
            elif row[0].startswith('Reported Claims'):
                row[0] = [{
                    'style': 'tbl head',
                    'background': '#eaf1dd',
                }, row[0]]
                for index, col in enumerate(row[1:]):
                    row[index + 1] = [{
                        'style': 'tbl strong r'
                    }, row[index + 1]]
            elif row[0].startswith('All Changes'):
                row[0] = [{
                    'style': 'tbl strong',
                    'background': '#eaf1dd',
                }, '\t{}'.format(row[0])]
            else:
                row[0] = [{
                    'style': 'tbl text',
                    'background': '#eaf1dd',
                }, '\t{}'.format(row[0])]

    # Insert 3-column span label
    rows.insert(0, [
        None,
        [{
            'colspan': 3,
            'style': 'tbl head c',
            'background': '#eaf1dd',
        }, 'Cumulative Lifetime Net Savings']
    ])

    # Add heading style to unit headers
    for index, cell in enumerate(rows[1]):
        if cell:
            rows[1][index] = [{
                'style': 'tbl head c',
                'background': '#eaf1dd',
            }, cell]
    return rows

def build_table6(data):
    rows = []
    for pa in ['PGE', 'SCE', 'SDGE', 'SCG']:
        rows += build_table6_pa(data, pa)
    return rows

def build_table6_pa(data, pa):
    rows = []
    rows.append([pa, None])
    for row in table6:
        if row.get('passthrough'):
            rows.append([row['index'], row['value']])
        elif not row['value']:
            rows.append([row['index'], None])
        else:
            rows.append([
                row['index'],
                eval(row['value'].format(pa=pa)),
            ])
    rows.append([None, None])
    return rows

def style_table6(raw):
    rows = deepcopy(raw)
    pa = rows[0][0]
    logging.info('Styling table 6: %s' % pa)

    # Remove first and last rows; they're intended to separate PAs in a list
    # of all 4 PA tables.
    del rows[0]
    del rows[-1]

    for row in rows:
        # Add heading format to first column.
        # Add bold format to appropriate data cells.
        if row[0]:
            # These cells get bold heading format.
            if (
                row[0] == 'Reported Claims'
                or row[0] == 'Savings Set to Zero'
                or row[0] == 'Other Savings Adjustments'
            ):
                row[0] = [{
                    'style': 'tbl head',
                    'background': '#eaf1dd',
                }, row[0]]
            # These are right-justified.
            elif (
                row[0] == 'No justification provided for early replacement'
                or row[0] == 'RUL less than one year'
            ):
                row[0] = [{
                    'style': 'tbl text 2',
                    'background': '#eaf1dd',
                }, '\t{}'.format(row[0])]
            # Everything else is indented.
            else:
                row[0] = [{
                    'style': 'tbl text 2',
                    'background': '#eaf1dd',
                }, row[0]]

            # Data cells are centered
            row[1] = [{
                'style': 'tbl text c',
            }, row[1]]
        else:
            # Column 2 heading ('Claims') is center bold.
            row[1] = [{
                'style': 'tbl head c',
                'background': '#eaf1dd',
            }, row[1]]

    return rows

def old_build_tablespecs():
    return [
        projects_and_claims_spec,
        claimed_vs_sampled_savings_spec,
        change_in_savings_spec,
    ]

def get_all_claim_data():
    logging.info('Retrieving all claim data')
    filename = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\8 PII\11 - Draft and Final Evaluation Reports\CIAC2018\Data\claimpop.csv'
    #return pd.read_excel(filename, sheet_name='claimpop')
    return pd.read_csv(filename)

def get_df(sfsession:ShareFileSession, tblsrc):
    """
    Load a dataframe using the passed info
    """
    path = tblsrc.source.source
    file_item = sfsession.get_io_version(path)
    if tblsrc.source.type == 'csv':
        file_item.io_data.seek(0)
        df = pd.read_csv(file_item.io_data)
    else:
        sheet = tblsrc.source.sheet
        headerrow = tblsrc.source.headerrow
        df = pd.read_excel(file_item.io_data, sheet_name=sheet, skiprows=headerrow-1)

    #filter the data if appropriate
    #TODO Implement this
    # if tblsrc.criteria:
            # criteria = 
    #     df = df[df[tblsrc]]

    return df

def build_tables(data, tablespecs):
    tables = []
    for spec in tablespecs:
        table = OldTable(data, spec)
        tables.append(table)
    return tables

def print_old_tables():
    tablespecs = old_build_tablespecs()
    tables = build_tables(tablespecs)
    for table in tables:
        preview_table(table.to_list())

def preview_table(table):
    """
    Prints a text-based preview of a table in list or tuple form.
    """
    print(tabulate(table, headers='firstrow', floatfmt=',.0f'))
    print()

def write_worksheet(ws, data, rStart=1, cStart=1):
    """
    This is legacy for D0. Now lives in tables.py
    Write a list of lists (or tuple of tuples) to cells in a worksheet.
    """
    for rowindex, row in enumerate(data):
        for colindex, col in enumerate(row):
            try:
                ws.cell(row=(rowindex + rStart), column=(colindex + cStart)).value = col
            except Exception as e:
                #know this happens for merged cells (except first cell of merged group)
                print(f'error putting {col} in cell {rowindex + rStart}, {colindex + cStart}. {e}') 
                pass

def export_data_D0():
    print('Loading from Design {}'.format(TEMPLATE_FILE_OLD))
    wb = load_workbook(TEMPLATE_FILE_OLD, data_only=False)
    tablespecs = old_build_tablespecs()
    all_claim_data = get_all_claim_data()
    tables = build_tables(all_claim_data, tablespecs)
    for table in tables:
        data = table.to_list()
        print(data)
        print()
        ws = wb['data_{}'.format(table.name)]
        write_worksheet(ws, data)
    write_worksheet(
        wb['data_5'],
        build_table5(
            get_table5_data(),
            all_claim_data,
        )
    )
    write_worksheet(
        wb['data_6'],
        build_table6(all_claim_data),
    )
    print('Saving to Report {}'.format(EXPORT_FILE))
    wb.save(EXPORT_FILE)

def export_data():
    """
    I believe this is old and not in use.
    """
    filename = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\GroupD-D11.01-CIAC 2018 Ex Post Evaluation -Design.xlsx'
    print('Loading from Design {}'.format(filename))
    #TODO change to io file handling
    wb = load_workbook(filename, data_only=False)
    specs = gettablespecs(filename)
    #filepath = r"Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\CIAC 2018 Tables and Figures - Design gh.xlsx"
    #specs = gettablespecs(filepath)

    #TODO this is just a data placeholder
    df = get_all_claim_data()

    keys = list(specs[0]['tables'].keys())
    for key in keys:
        print('Creating table {}'.format(key))
        logging.info('Creating table {}'.format(key))
        table_spec = specs[0]['tables'][key]
        #TODO get df's from the spec
        table = Table.from_spec(table_spec, df=df)
        #TODO There's a problem just using cell data because it has formatted the numbers with commas
        # also may want to limit it to just the data section
        data = table.to_data_list(table.datastart['row'],table.datastart['col'])
        print(data)
        print()
        ws = wb[table.name]
        #ws = wb['PAFirstElecSavings']

        write_worksheet(ws, data, rStart=table.datastart['row'], cStart=table.datastart['col'])

    #TODO get real output location/name
    outputfolder = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\output'
    outputfile = outputfolder + r'\reportspreadsheet.xlsx'
    print('Saving to Report {}'.format(outputfile))
    
    #TODO add Io handling
    wb.save(outputfile)

def export_to_word():
    # tablespecs = old_build_tablespecs()
    all_claim_data = get_all_claim_data()
    # old_tables = build_tables(all_claim_data, tablespecs)
    # tables = []
    # for table in old_tables:
    #     data = table.to_list()
    #     tables.append(Table.from_list(data, table.title))

    for pa in ['PGE', 'SCE', 'SDGE', 'SCG']:
        table5_raw = build_table5_pa(get_table5_data(), all_claim_data, pa)
        table5_styled = style_table5(table5_raw)
        table = Table.from_list(
            table5_styled,
            'Cumulative Changes Due to Database'
            ' and Sample Review for {}'.format(PA_MAP[pa]),
        )
        table.to_word(WORD_TABLE_TEMPLATE, 'tmp/table5_{}.docx'.format(pa))
        # table.col_widths = [2.25, 1.25, 1.25, 1.25]
        # tables.append(table)
    # return tables

def transferdatadefs(sfsession, filename, sheetname, wb, ):
    """
    Copy sheets from the datadefs workbook to the report workbook
    if sheet does not exist will create it, otherwise it will just update the contents starting at the table start position
    """

    #open the workbook
    file_item = sfsession.get_io_version(filename)
    if not file_item:
        return False
    wb_src = load_workbook(filename=file_item.io_data, data_only=True)
    if not wb_src:
        return False
    #grab the list
    try:
        df = pd.read_excel(file_item.io_data, sheetname)
    except:
        print(f'problem getting list for filename')
    # sheetlist = df[sheetname].to_list()
    table_start_markers = [params.RPT_TABLE_START_MARKER, params.RPT_TABLE_START_RANGE_MARKER]
    #loop through the list 
    # for sheet in sheetlist:
    for _, row in df.iterrows():
        sheet = row[sheetname]
        desc = row[1]
        #check for sheet in passed wb
        if sheet not in wb.sheetnames:
        #create it if it's not there (including add table start marker)
            ws_dest = wb.create_sheet(sheet) # can add numeric position if want. Use 0 to put at beginning or leave blank for the end
            ws_dest.title = sheet
            ws_dest.cell(1,2).value = desc
            ws_dest.cell(3,1).value = params.RPT_TABLE_START_MARKER
        else:
            ws_dest = wb[sheet]

        #get tablestart offset
        maxdeep = 20
        r_offset = None
        for i in range(1, maxdeep):
            if r_offset:
                break
            for j in range(1,maxdeep):
                if ws_dest.cell(i,j).value in table_start_markers:
                    r_offset = i - 1
                    c_offset = j - 1
                    break
        if not r_offset: #in case the marker has been written over, try to find where the data starts
            for i in range(2, maxdeep):
                if r_offset:
                    break
                for j in range(1,maxdeep):
                    if ws_dest.cell(i,j).value is not None:
                        r_offset = i - 1
                        c_offset = j - 1
                        break
        if not r_offset:
            r_offset = 2
            c_offset = 0

        #do a cell by cell data transfer
        ws_src = wb_src[sheet]
        #get the bounds
        mr = ws_src.max_row
        mc = ws_src.max_column
        
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                # reading cell value from source excel file
                c = ws_src.cell(row = i, column = j)
                # writing the read value to destination excel file
                ws_dest.cell(row = i + r_offset, column = j + c_offset).value = c.value
                #Font, PatternFill, Border, Alignment
                ws_dest.cell(row = i + r_offset, column = j + c_offset).font = copy(c.font)
                ws_dest.cell(row = i + r_offset, column = j + c_offset).fill = copy(c.fill)
                ws_dest.cell(row = i + r_offset, column = j + c_offset).number_format = copy(c.number_format)
                ws_dest.cell(row = i + r_offset, column = j + c_offset).alignment = copy(c.alignment)
                ws_dest.cell(row = i + r_offset, column = j + c_offset).border = copy(c.border)

def run_google_download_driver(driverpath):
    """
    open the driver file and download specified files
    """

    #open driver
    df = excel_to_df(driverpath)
    df = df[df['active']=='y']

    for _, row in df.iterrows():
        download_drive_file(row['fileid'], row['dest'], row['exporttype'])


def run_generate_tables(driverpath):
    """
    generate the tables based on the parametes specified in the passed file
    """
    sfsession = ShareFileSession(SHAREFILE_OPTIONS)
    logfile = StringIO()
    create_logfile(logfile, level=logging.NOTSET)
    #set below to warning, because can't get it to log otherwise. despite setting levels everywhere
    sys = platform.uname()
    logging.warning(f'Running on {sys.processor}, node {sys.node}, version:{sys.version} : {str(datetime.now())}')
    logging.warning('INFO - Beginning session from main')
    logroot = 'generatetables_'
    logextension = datetime.now().strftime('%m-%d-%Y_%H-%M')
    sfsession.upload_file(params.LOG_PATH_ID, logroot + logextension + '.log', logfile)

    #open driver file
    df = excel_to_df(driverpath)
    df = df[df['active'].str.lower()=='y']

    for _, row in df.iterrows():
        #get busy        
        filename = row['sourcefile']
        statusfilter = row['statusfilter']
        generateword = row['word']
        generateexcel = row['excel']
        generateplots = row['plots']
        outputfolder = row['outputfolder']
        testing = row['testing']
        postprocess = row['postprocess']
        googlesource = row['googlesource']
        
        #download design file if on google
        if isinstance(googlesource, str): #blanks come in as nan floats            
            download_drive_file(googlesource, filename)

        newname = os.path.basename(filename).replace('Design', 'Report' + '_' + statusfilter + '_ ' +logextension)
        outputfile = os.path.join(outputfolder, newname)
    
        file_item = sfsession.get_io_version(filename)
        if not file_item:
            logging.critical(f'no report template file {filename}')
        else:            
            if generateexcel:
                wb = load_workbook(file_item.io_data, data_only=False)
            #can't pass the workbook because it's opened as data_only=False and spec needs the values
            specs = gettablespecs(sfsession, file_item.io_data, statusfilter)
            
            if generateexcel or generateword or postprocess:
                keys = list(specs[0]['tables'].keys())
                if postprocess:
                    wrdapp = openWord()              
                for key in keys:
                    print('Creating table {}'.format(key))
                    logging.warning('Creating table {}'.format(key))
                    table_spec = specs[0]['tables'][key]
                    #df = load_df(table_spec.source, sourcesfile)
                    #TODO load the df listed in spec. maybe into a list to reduce reloading time if memory space is enough
                    
                    df = get_df(sfsession,table_spec)
                    table = Table.from_spec(table_spec, df=df)

                    if generateword:
                        if table.cells:
                            tmpl_item = sfsession.get_io_version(WORD_TABLE_TEMPLATE)
                            if not tmpl_item:
                                msg = f'cannot get word template from {WORD_TABLE_TEMPLATE}'
                                print(msg)
                                logging.critical(msg)
                            else:
                                table.to_word(sfsession, tmpl_item.io_data, '{}\\{}.docx'.format(outputfolder, key))
                    if generateexcel:
                        table.to_excel(wb)
                    if postprocess:
                        try:
                            table.post_process(wrdapp, '{}\\{}.docx'.format(outputfolder, key))
                            # print(f'dont post processing {key}')
                        except Exception as e:
                            msg = f'Problem postprocessing {key}: {e}'
                            print(msg)
                            logging.critical(msg)
                    msg = f'done with table {key}'
                    print(msg)
                    logging.warning(msg)
                    try:
                        sfsession.upload_file(params.LOG_PATH_ID, logroot + logextension + '.log', logfile)
                    except:
                        print('Drat! Failed to upload log file after generating tables')

            if generateexcel:
                formatws(wb) #added by SM to format sheets
                #Transfer datadefs sheets
                #TODO pull this info from teh design file so it's not specified here.
                # transferdatadefs(sfsession, params.CIAC_2018_DATA_DEF_FILE, 'exportsheets', wb)                
                file_io = BytesIO()
                wb.save(file_io)        
                #upload to sf
                output_item = sfsession.get_item_by_local_favorites_path(outputfile)
                if output_item:
                    wkb_folderID = output_item.data['Parent']['Id']
                else:
                    folder_item = sfsession.get_item_by_local_favorites_path(outputfolder)
                    if folder_item:
                        wkb_folderID = folder_item.id
                try:
                    sfsession.upload_file(wkb_folderID, newname, file_io)
                except:
                    logging.warning(f'unable to save {newname}')
                file_io = None
            
            if generateplots:
                keys = list(specs[1]['figures'].keys())
                logging.warning('Starting figures')
                for key in keys:
                    spec = specs[1]['figures'][key]
                    callstring = f'CIAC2018.{spec.source.source} ({spec.captions}, None, printdata=True)'                
                    print(f'callstring is:{callstring}')
                    results = eval(callstring)
                    logging.warning(f'done with figure {key}')
                    try:
                        sfsession.upload_file(params.LOG_PATH_ID, logroot + logextension + '.log', logfile)
                    except:
                        print('Failed to upload log file after figures')
            if testing:
                #Print the fields used
                fieldlist = dict()
                for _,table in specs[0]['tables'].items():
                    if table.source.name in fieldlist:
                        #TODO fix below it is not preventing dupilcate entries as expected
                        fieldlist[table.source.name] = fieldlist[table.source.name] + list(set(table.uniquefields) - set(fieldlist[table.source.name]))
                        # for field in table.uniquefields:
                        #     if field not in fieldlist[table.source.name]:
                        #         fieldlist[table.source.name] = fieldlist[table.source.name] + [field]
                    else:
                        fieldlist[table.source.name] = list(set(table.uniquefields))
                    # print(fieldlist[table.source.name])
                df=pd.DataFrame.from_dict(fieldlist,orient='index').transpose()
                filename = 'usedfields.csv'
                outputfile = os.path.join(outputfolder, filename)
                df.to_csv(outputfile)

        logging.warning('INFO - Process complete')
        sfsession.upload_file(params.LOG_PATH_ID, logroot + logextension + '.log', logfile)
    return True

if __name__ == '__main__':
  
    run_generate_tables(r"Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\generate_tables_driver.xlsx")
    # run_generate_tables(r"Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\generate_tables_driver_JSL.xlsx")
