#read report table spec from excel
import logging
import os.path
import operator
import copy

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
import pandas as pd

from cpuc.sharefileapi import ShareFileSession
from cpuc.sharefileapi import SHAREFILE_OPTIONS

from cpuc.mylogging import create_logfile
import cpuc.params as params
from cpuc.utility import rgb_to_hex
from cpuc.workbookfunctions import openworkbook
from cpuc.workbookfunctions import convertwstodf

headermrk = 'h'
markers = [headermrk, 'di', 'd', 'c'] #Header, data in place, data, calculation
passthrumarkers = [headermrk,'di'] #these indicate that the contents should be passed through without change

table_start_markers = [params.RPT_TABLE_START_MARKER, params.RPT_TABLE_START_RANGE_MARKER]
start_and_header = table_start_markers + [headermrk]

class TableSpec():
    """
    Class for containing all the specification information for a report table
    """
    def __init__(self, **kwargs):
        self.type = 'table'
        self.source = None
        self.name = None
        self.dimensions = None
        self.orientation = None
        self.style = 'Normal' #dense is the other option
        self.autofit = 'content' #window is the other option
        self.cells = None
        self.fields = None
        self.uniquefields = None
        self.comments = None
        self.sheet = None
        self.datastart = None #the first nonheader cell
        self.tablestart = None #the location of the table start marker for offsetting the excel version
    
    '''
        dimensions is a dict containing the number of rows and columns in the table:
            {
                'rows': 10,
                'columns': 4,
            }

        cells is a list of lists. Each top-level list item is a row in the table, and
        each sublist item is a cell.
            [
                [
                    ['cell1'],
                    ['cell2'],
                    ...
                ],
                ...
            ]

        comments is a dict of comments. The keys are the comment, the value is the comment number
            {
                'comment blah': 1, 
                'blah blah': 2
            }

    '''
    def getdataspec(self, ws):
        """
        Pull together the information required to put data into the worksheet
        """
        #for destination report workbook
        #gather the field names only from the sheet
        #get the start location = datastart cell

        #get table marker
        self.tablestart = gettablemarkerloc(ws)
        self.sheet = ws.title
        self.type = 'data'

        #get the fields by going down from start until a blank
        row = self.tablestart['row']
        col = self.tablestart['col']        
        offset = 1
        fields = []
        while ws.cell(row + offset, col).value is not None:
            fields.append(ws.cell(row + offset, col).value)
            offset += 1

        if len(fields) == 1: #it set up column wise
            offset = 1
            row +=1
            while ws.cell(row, col+ offset).value is not None:
                fields.append(ws.cell(row , col+ offset).value)
                offset += 1
        self.fields = fields        

    def get_unique_fields(self):
        """
        Print the fields that are used in each table
        """
        fieldlist = []
        if self.fields:           
            for field in self.fields:
                fields = field.split(',')            
                for field in fields:
                    if field.strip() not in fieldlist:
                        fieldlist.extend(field.strip())
        elif self.cells:
            for row in self.cells:
                for cell in row:
                    if cell[0]['fields']:
                        fields = cell[0]['fields']
                        # for field in cell[0]['fields']:
                        fields = fields.split(',')
                        for field in fields:
                            if field.strip() not in fieldlist:
                                fieldlist.append(field.strip())


        self.uniquefields = fieldlist

# class FigureSpec():


class Pfield(object):
    """ Class to specify parameter field
    """
    def __init__(self, column, field):
        self.column = column
        self.field = field

class Cellprops(object):
    """ Class to hold the various cell properties
    """
    def __init__(self):
        self.type = None #header, data, def
        self.colid = None #for connecting column specific fields
        self.text = None
        self.background = None
        self.alignment = None
        self.comment = None
        self.commentid = None
        self.bold = None
        self.mergespan = None
        self.format = None
        self.decimals = None
        self.percent = None
        self.fields = None
        self.filters = None
        self.colfields = None
        self.colfilters = None
        self.colcalc = None #for createing calculated columns, not implemented yet
        self.transforms = None
        self.reducer = None
        self.calc = None
        self.calcfield = None
        self.row = None
        self.col = None
        self.wraptext = None
        self.colwidth = None

class SourceDef:
    """
    Class to hold all the information necessary to create a dataframe from a datasource and keep track of it
    """
    def __init__(self):
        self.name = None
        self.type = 'csv'
        self.sheet = None
        self.headerrow = None
        self.source = None #this should be either a path to the object or an io version of the object
        self.indexfield = None
        self.fkeyrelationshsip = None
        self.criteria = None
        self.renameindex = None
    
    # @classmethod
    def get_details(self, sfsession, wb, sheet:str, srccol:str, srcname:str, loccol:str, shtcol:str, tablename:str, headerrow=1, srcmap=None):
        """
        populate the properties for the source
        It's a 2 stage lookup. 
            1. lookup the location of the source def file
            2. lookup source file details
            assumes source map has column names: sourcename, type, location, IndexField, ForeignKeyRelationship, criteria, sheet, startrow, rename_index
            There is a default attribute source map, but can pass a different one. Form is a dict with attributes as keys

        """
        self.name = tablename
        lkptbl_df = convertwstodf(wb[sheet], headerrow=headerrow)
        #pull def location from src table
        lkptbl_df = lkptbl_df[lkptbl_df[srccol] == srcname]

        path = lkptbl_df.loc[: , loccol].values[0]
        sheet = lkptbl_df.loc[: , shtcol].values[0]
        
        src_item = sfsession.get_io_version(path)
        try: #open source def file            
            lkptbl_df = pd.read_excel(src_item.io_data, sheet_name=sheet) #assumes headeris on first row
        except:
            try:
                lkptbl_df = pd.read_csv(src_item.io_data)
            except:
                print('cannot open source map path')
                return self

        #pull def specifics from src table
        try:
            srctbl_df = lkptbl_df[lkptbl_df['sourcename'] == tablename]
        except:
            print(f'column heading do not match for file {path}')
            return
        if len(srctbl_df.index) != 1:
            print(f'cannot get details for {tablename}')
            return self

        #map out the attributes to the colnames
        if not srcmap:
            srcmap = {
                'type':'type',
                'source': 'location',
                'indexfield': 'indexfield',
                'fkeyrelationshsip': 'foreignkeyrelationship',
                'criteria': 'criteria',
                'sheet': 'sheet',
                'headerrow': 'startrow',
                'renameindex': 'rename_index',
            }

        for k,v in srcmap.items():
            col = [x for x in srctbl_df.columns if x.lower() == v]
            setattr(self, k, srctbl_df[col].values[0][0])
            
        # print('done')
        return self


def gettablespecs(sfsession, filepath, statusfilter='Test'):
    """
    Load the various table specs from the passed file as defined on the Captions sheet
    Inputs - Filepath or workbook, statusfilter to control which tables
    Returns a list of tablespec objects (and future figure spec) 
    """
    #open workbook
    #assumes there is a captions sheet which will drive it
    sheet = 'Captions'
    srcsheet = 'SourceDefs'    

    #TODO bring in def as a dataframe instead of messing with it as a sheet.
    if isinstance(filepath, str):
        file_item = sfsession.get_io_version(filepath)
        if file_item:
            filepath = file_item.io_data
    try:
        wb = openworkbook(filepath, values=True)
    except:    
        wb = filepath

    if not wb:
        print('problem loading workbook. Quitting')
        return False
    try:
        ws = wb[sheet]
    except:
        print('Captions sheet missing')
        return False
    try:
        #ws_src = None
        ws_src = wb[srcsheet]
    except:
        print('Source defs sheet missing')
        return False
    #cycle through the captions listed if type is table
    #for row in ws.iter_rows(row_offset=1):
    #get critical col numbers
    for i in range(1,ws.max_column + 1):
        if ws.cell(1,i).value is None:
            continue
        #print(f'text is {ws.cell(1,i)}')
        if ws.cell(1,i).value.lower() == 'type':
            typecol = i
        elif ws.cell(1,i).value.lower() == 'text':  #for old version
            captioncol = i
        elif ws.cell(1,i).value.lower() == 'caption': #for new version
            captioncol = i
        elif ws.cell(1,i).value.lower() == 'sheet':
            sheetcol = i
        elif ws.cell(1,i).value.lower() == 'status':
            statuscol = i
        elif ws.cell(1,i).value.lower() == 'source':
            sourcecol = i
        elif ws.cell(1,i).value.lower() == 'destination':
            destcol = i
        elif ws.cell(1,i).value.lower() == 'style':
            stylecol = i
        elif ws.cell(1,i).value.lower() == 'fit':
            fitcol = i

    parameters = []
    #not using this section after D0 (I don't think)
    '''
    if 'sheetcol' in locals():
        #to pull filter column data. PA in the D0 example
        for i in range(sheetcol + 2, ws.max_column + 1): #to skip extra column
            param = Pfield(i, ws.cell(1,i).value)
            #parameters.append({pcol:i, field:ws.cell(1,i)})
            parameters.append(param)
    '''
    
    tablespecs = {}
    figurespecs = {}
    for i in range(1,ws.max_row +1):
        #print(f'processing row {i} as sheet {ws.cell(i,sheetcol).value}')
        #test below breaks on D0 so no longer compatible :(
        
        if ws.cell(i,destcol).value is not None and ws.cell(i,destcol).value.lower() == 'report workbook' and ws.cell(i,statuscol).value is not None and statusfilter.lower() in ws.cell(i,statuscol).value.lower():
            ws_object = wb[ws.cell(i,sheetcol).value]
            tblspec = TableSpec()
            tblspec.getdataspec(ws_object)                        
            tblspec.source = SourceDef().get_details(sfsession, wb=ws.parent, sheet='SourceDefs', srccol='sourcename', srcname='sourcedef', loccol='location', shtcol='sheet', tablename=ws.cell(i,sourcecol).value )
            tblspec.name = ws.cell(i,captioncol).value
            tblspec.get_unique_fields()
            tablespecs[tblspec.name] = tblspec
        elif ws.cell(i,destcol).value is not None and ws.cell(i,typecol).value.lower() == 'table' and ws.cell(i,statuscol).value is not None and statusfilter.lower()  in ws.cell(i,statuscol).value.lower():
            lkpsheet = ws.cell(i,sheetcol).value
            if lkpsheet not in wb.sheetnames:
                print(f'yo! sheet missing:{lkpsheet}')
                continue
            ws_object = wb[ws.cell(i,sheetcol).value]
            #ws_object = wb[ws.cell(i,captioncol).hyperlink.location.split('!')[0].replace("'", '')]            
            sfilter = ''
            for item in parameters:
                if ws.cell(i,item.column).value:
                    if ws.cell(i,item.column).value.isnumeric():
                        sfilter += f'{item.field} == {ws.cell(i,item.column).value} and '
                    else:
                        sfilter += f'{item.field} == \'{ws.cell(i,item.column).value}\' and '
            if sfilter == '':
                sfilter = None
            elif sfilter[-4:] == 'and ':
                sfilter = sfilter[:-5]

            tblspec = createtablespec(ws_object, sfilter)

            if not tblspec:
                print(f'problem getting tablespec for {ws_object.title}')
                continue 
            tblspec.source = SourceDef().get_details(sfsession, wb=ws.parent, sheet='SourceDefs', srccol='sourcename', srcname='sourcedef', loccol='location', shtcol='sheet', tablename=ws.cell(i,sourcecol).value )            
            tblspec.name = ws.cell(i,captioncol).value
            tblspec.style = ws.cell(i,stylecol).value
            tblspec.autofit = ws.cell(i,fitcol).value
            tblspec.get_unique_fields()
            tablespecs[tblspec.name] = tblspec
        elif ws.cell(i,destcol).value is not None and ws.cell(i,typecol).value.lower() == 'figure':
            ws_object = wb[ws.cell(i,captioncol).hyperlink.location.split('!')[0].replace("'", '')]
            figurespecs[ws.cell(i,captioncol).value] = createfigurespec(ws_object) #this call currently doesn't do anythign because the procedure is just a pass

    #possibly only temporary processing of figures
    #TODO change to io verion
    df_captions = pd.read_excel(filepath, sheet)       
    df_figures = df_captions[df_captions['Type'].str.contains('Plot') & df_captions['Status'].str.contains(statusfilter)].groupby('Source')
    
    for source, data in df_figures:
        figurespec = TableSpec()
        figurespec.captions = [x for x in data['Caption']]
        figurespec.source = SourceDef().get_details(sfsession, wb=ws.parent, sheet='SourceDefs', srccol='sourcename', srcname='sourcedef', loccol='location', shtcol='sheet', tablename=source)
        figurespecs[source] = figurespec

    specs = []
    specs.append({'tables':tablespecs})
    specs.append({'figures':figurespecs})
    return specs


def createtablespec(ws, parameters = None) -> TableSpec:
    #TODO move this stuff into the tablaspec class
    tablebounds = gettableboundaries(ws)
    tabledef = []
    fieldprops = []

    #pull together defs so they can be added to each cell as appropriate
    if not tablebounds:
        return None

    for i in range(tablebounds['rows'][0]+1,ws.max_row + 1):
        for j in range(tablebounds['cols'][0]+1,ws.max_column + 1):
            tblcelldefs = Cellprops()
            #print(f'for row {i} col {j} the value is {ws.cell(i,j).value}')
            if (ws.cell(i,j).value is not None
                    and ws.cell(tablebounds['rows'][0], j).value is not None
                    and ws.cell(i, tablebounds['cols'][0]).value is not None
                    and (ws.cell(tablebounds['rows'][0], j).value not in markers
                        or ws.cell(i, tablebounds['cols'][0]).value not in markers)
                    and ws.cell(i,j).value not in table_start_markers):
                #print(f'adding cell def for row {i} column {j}')

                tblcelldefs = fillcelldefs(ws.cell(i,j), tblcelldefs, ws.cell(i, tablebounds['cols'][0]).value, ws.cell(tablebounds['rows'][0], j).value)
                fieldprops.append(tblcelldefs)
            else:
                #print(f'skipping cell def for row {i} column {j}')
                pass
    
    tablecomments = dict()    
    for i in range(tablebounds['rows'][0]+1,tablebounds['rows'][1]+1):
        for j in range(tablebounds['cols'][0]+1,tablebounds['cols'][1]+1):
            if ws.cell(i,j).value is not None:
                ocell = ws.cell(i,j)
                tblcell = Cellprops()
                tblcell.row = i
                tblcell.col = j
                tblcell.alignment = ocell.alignment.horizontal
                tblcell.bold = ocell.font.b
                tblcell.background = translatecolortohex(ws, ocell)
                tblcell.comment = getcommenttext(ocell)                
                tblcell.decimals = numberofdecimals(ocell)
                tblcell.format = ocell.number_format
                tblcell.mergespan = isMergedCellStart(ocell)
                tblcell.percent = '%' in ocell.number_format
                tblcell.wraptext = ocell.alignment.wrapText
                #if ws.cell(tablebounds['rows'][0], j).value == 'h' or ws.cell(i, tablebounds['cols'][0]).value == 'h':
                if ws.cell(tablebounds['rows'][0], j).value in passthrumarkers or ws.cell(i, tablebounds['cols'][0]).value in passthrumarkers:
                    tblcell.text = str(ocell.value)
                    #tblcell.type = 'h'
                    tblcell.type = ws.cell(i, tablebounds['cols'][0]).value
                elif ws.cell(tablebounds['rows'][0], j).value == markers[2] and ws.cell(i, tablebounds['cols'][0]).value == markers[2]:
                    tblcell.type = markers[2]
                elif ws.cell(tablebounds['rows'][0], j).value is not None and ws.cell(i, tablebounds['cols'][0]).value is not None:
                    print('uh oh, should not get here')
                    #tblcell = fillcelldefs(ws.cell, tblcell, ws.cell(tablebounds['rows'][0], j).value, ws.cell(i, tablebounds['cols'][0]).value)
                else:
                    #tblcell.field =
                    print('hmm, how did I get here')
                    pass

                if tblcell.type == markers[2]: #data
                    tblcell = addcelldefs(tblcell, fieldprops)
                    #TODO Add check for field count match for transform and reducers?
                    if parameters is not None:
                        if tblcell.filters is None:
                            tblcell.filters = parameters
                        else:
                            tblcell.filters += ' and ' + parameters

                if tblcell.comment is not False:
                    #check to see if this comment is already in the table                    
                    if tblcell.comment in tablecomments:
                        #get comment number
                        tblcell.commentid = tablecomments[tblcell.comment]                        
                    else:   #comment needs to get added to the list
                        tblcell.commentid = len(tablecomments)+1
                        tablecomments[tblcell.comment] = tblcell.commentid #len(tablecomments)+1
                        # tblcell.comment = f'{tablecomments[tblcell.comment]} tmp for testing {tblcell.comment}'
                        # tblcell.comment = tablecomments[tblcell.comment]
                        


                #adjust row and col to be relative to table not sheet
                tblcell.row = tblcell.row - tablebounds['rows'][0]
                tblcell.col = tblcell.col - tablebounds['cols'][0]

                tabledef.append(tblcell)

    #Fix single table comment number
    if len(tablecomments) == 1: #spin back through to change the 1 to an asterisk
        for cell in tabledef:
            if cell.commentid and cell.commentid == 1:
                cell.commentid = '*'
                # cell.comment = cell.comment.replace('1', '*', 1)
        tablecomments = {k:'*' for k, v in tablecomments.items() if v==1}

    #get the cells for the table
    olist = createtablelist(tabledef)
    #combine dimensions and table def
    columns = tablebounds['cols'][1] -tablebounds['cols'][0]
    #arbitrary way to set orientation. Add in variable on sheet to set this
    # if columns > 7:
        # orientation = 'landscape'
    # else:
    orientation = 'portrait'

    otbl = TableSpec()
    otbl.dimensions = {'rows': tablebounds['rows'][1]- tablebounds['rows'][0], 'columns': columns}
    otbl.orientation =  orientation
    otbl.cells =  olist
    otbl.comments = tablecomments

    #Get source

    #get datastart
    otbl.datastart = gettabledatastart(ws)

    #get table marker
    otbl.tablestart = gettablemarkerloc(ws)

    #Get sheetname
    otbl.sheet = ws.title

    return otbl


def createfigurespec(worksheet, params = None):
    """

    """


    return True

def isMergedCellStart(cell):

    for item in sorted(cell.parent.merged_cells.ranges):
        if item.left[0][0] == cell.row and item.left[0][1] == cell.column:
            # return item.size['columns']
            return item.size
    return False

def numberofdecimals(cell):
    #'_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
    #'0.0'
    if '.' in cell.number_format:
        tmp = cell.number_format.split('.')[1]
        #print(f'tmp is {tmp}')
        counter = 0
        for c in tmp:
            if c != '0':
                break
            counter = counter + 1
        return counter
    return False

def getcommenttext(cell):
    if cell.comment is None:
        return False
    else:
        #drop threaded comments
        if '[Threaded comment]' in cell.comment.content:
            return False
        else:
            return cell.comment.content

def gettableboundaries(ws):
    startcell = None
    for i in range(1, ws.max_column +1):
        for j in range(1, ws.max_row + 1):
            # if ws.cell(j, i).value == RPT_TABLE_START_MARKER or ws.cell(j, i).value == RPT_TABLE_START_RNAGE_MARKER:
            if ws.cell(j, i).value in table_start_markers:
                startcell = ws.cell(j, i)
                break
    if startcell is None:
        return False
    rows = startcell.row
    cols = startcell.column

    for i in range(startcell.row , ws.max_row +1):
        if ws.cell(i,startcell.column ).value is not None and ws.cell(i,startcell.column ).value in markers:
            rows += 1
    for i in range(startcell.column, ws.max_column +1):
        if ws.cell(startcell.row ,i).value is not None and  ws.cell(startcell.row ,i).value in markers:
            cols += 1
    return {'rows':[startcell.row, rows],'cols':[startcell.column, cols]}

def gettabledatastart(ws):
    #returns the row annd col for the first
    startcell = None
    for i in range(1, ws.max_column +1):
        for j in range(1, ws.max_row + 1):
            if ws.cell(j, i).value in table_start_markers:
                startcell = ws.cell(j, i)
                break
    if startcell is None:
        return False
    rows = startcell.row
    cols = startcell.column

    for i in range(startcell.row , ws.max_row +1):
        if ws.cell(i,startcell.column).value is not None and (ws.cell(i,startcell.column).value in start_and_header):
            rows += 1
        else:
            break
    for i in range(startcell.column, ws.max_column +1):
        if ws.cell(startcell.row ,i).value is not None and (ws.cell(startcell.row ,i).value in start_and_header):
            cols += 1
        else:
            break
    
    return {'row':rows,'col':cols}

def gettablemarkerloc(ws):
    #returns the row and col for the table marker    
    for i in range(1, ws.max_column +1):
        for j in range(1, ws.max_row + 1):
            if ws.cell(j, i).value in table_start_markers:
                return {'row':j, 'col': i}               
    return {'row':j, 'col': i}

def fillcelldefs(cell, tblcell, rowhdr, colhdr):
    #row header is the header for rows of defs under the table
    fieldsep = '_'
    #fieldprefix = 'field_'
    colidvar = 'colid'

    tblcell.row = cell.row
    tblcell.col = cell.column
    #populate the attributes that match the row or col headers around the table
    for attr, value in tblcell.__dict__.items():
        if attr == rowhdr:
            setattr(tblcell, attr, cell.value)
        if attr == colhdr:
            setattr(tblcell, attr, cell.value)

    
    if rowhdr == colidvar:  #add the column id
        tblcell.colid = cell.value
    elif fieldsep in colhdr: #add the column level specifications
        try:
            setattr(tblcell, colhdr.split('_')[0], cell.value)
        except:
            print(f"attribute missing: {colhdr.split('_')[0]}")
        #tblcell.field = cell.value
        tblcell.colid = colhdr.split('_')[1]

    return tblcell

def addcelldefs(tblcell, tbldefs):
    #look for row and col match or row and colid match
    #copy fields over
    #assign colid if present
    #hascolid = False

    for cell in tbldefs:
        if cell.colid is not None:
            if cell.col == tblcell.col:
                tblcell.colid = cell.colid
                #hascolid = True
                break

    noupdate = ['row', 'col']
    #print(f'row is {tblcell.row} and col is {tblcell.col}')
    for cell in tbldefs:
        #add column defs
        if cell.col == tblcell.col:
            for attr, value in cell.__dict__.items():
                #print(attr, value)
                if value is not None and attr not in noupdate:
                    setattr(tblcell, attr, value)

        # Make sure reducer is set on data cells, not just the row header
        #this may need to add another criteria like colid check
        if cell.row == tblcell.row and cell.reducer is not None:
            tblcell.reducer = cell.reducer
            #print(tblcell)

    for cell in tbldefs:
        #if cell.row == tblcell.row and (not hascolid or cell.colid == tblcell.colid):
        if cell.row == tblcell.row and cell.colid == tblcell.colid:
            for attr, value in cell.__dict__.items():
                #print(attr, value)
                if value is not None and attr not in noupdate:
                    setattr(tblcell, attr, value)

    return tblcell

def createtablelist(tbl):
    #create the list format for the word table creation
    # It's a two element list.
    #   Item 1 is a dictionary of metadata about the cell
    #   Item 2 is the value. It will be blank for empty cells, text for text stuff, and null for values that need to be calculated or looked up
    tbllist = []

    tbl.sort(key = operator.attrgetter('row', 'col'))
    #tbl = sorted(tbl, key = lambda row:cell.row)
    #tbl = sorted(tbl, key = lambda col:cell.col)
    rownum = 0
    for cell in tbl:
        tblmeta = {}
        #print(f'creating list for row {cell.row} and column {cell.col}')
        if cell.row != rownum:
            if rownum != 0:
                tbllist.append(row)
            row = []

        for attr, value in cell.__dict__.items():
            #setattr(tblmeta, attr, value)
            tblmeta[attr] = value

        #handle special attributes
        style = 'tbl'
        if cell.type == markers[0]:
            style += ' head'
        elif cell.bold == True:
            style += ' strong'
        else:
            style += ' text'

        if cell.alignment is None:
            pass
        elif 'right' in cell.alignment:
            style += ' r'
        elif 'center' in cell.alignment:
            style += ' c'
        #setattr(tblmeta, 'style', style)
        tblmeta['style'] = style

        if cell.decimals is None or cell.decimals == False:
            #setattr(tblmeta, 'decimals', 0)
            tblmeta['decimals'] = 0

        if cell.text is not None:
            value = cell.text
        else:
            value = None

        if cell.text is None and cell.type is None:
            row.append([None])
        else:
            row.append([tblmeta, value])

        rownum = cell.row

    tbllist.append(row)

    return tbllist

def translatecolortohex(worksheet, cell):
    #look up on colortrans sheet
    lkp_sheet = 'automation'
    defcolor = '#eaf1dd' #green
    # defcolor = '#FFFFFF' #white
    theme = cell.fill.fgColor.theme
    if not isinstance(theme, int):
        rawcolor = cell.fill.fgColor.rgb
        color = rgb_to_hex(rawcolor)
        # try:
        #     color = int(rawcolor)
        #     if len(color) == 9:
        #         color = rgb_to_hex(color)
        # except:
        #     color = rawcolor[2:]
        return color
    tint = round(cell.fill.fgColor.tint,1)
    try:
        ws = worksheet.parent[lkp_sheet]
    except:
        return defcolor

    for i in range(1,ws.max_row +1):
        if ws.cell(i,1).value == theme and ws.cell(i,2).value == tint:
            return ws.cell(i,3).value

    return False
    #themes = list(range(1,11)
    #tints = [0.7999816888943144, 0.5999938962981048]

def get_source_details():
    """
    Get the info to fill the SourceDef class
    """


if __name__ == '__main__':
    #spec = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\00 - 2017 Evaluation\Report\Drafts\2019-5-15\D0 Report Figures and Tables - Design.xlsx'
    #d0spec = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\CIAC 2018 Tables and Figures - Design gh.xlsx'
    #ciac2018
    sfsession = ShareFileSession(SHAREFILE_OPTIONS)
    # spec = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\Copy of GroupD-D11.01-CIAC 2018 Ex Post Evaluation -Design_gh.xlsx'
    spec = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\GroupD-D11.01-CIAC 2018 Ex Post Evaluation -Design_ginatest.xlsx'
    filepath = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\Design_exceltest.xlsx'
    # filepath = r'Z:\Favorites\CPUC10 (Group D - Custom EM&V)\4 Deliverables\11 - Draft and Final Evaluation Reports\CIAC 2018\Design\GroupD-D11.01-CIAC 2018 Ex Post Evaluation -Design.xlsx'
    specs = gettablespecs(sfsession,filepath, 'Test')
    print(specs)
