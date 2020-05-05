#import openpyxl
from openpyxl import load_workbook


filename = 'P:\\Projects\\CPUC10 (Group D - Custom EM&V)\\4 Deliverables\\21 - Other Contractor Coordination\\ginatemp\\Draft_Gross_ExPost_Enhanced.xlsx'
wb = load_workbook(filename)
print ('sheets:' + wb.sheetnames)

names = wb.defined_names
print('names: ' + names)
#my_range = wb.defined_names['my_range']
# if this contains a range of cells then the destinations attribute is not None
#dests = names[0].destinations # returns a generator of (worksheet title, cell range) tuples
#print ('dest: ' + dests)