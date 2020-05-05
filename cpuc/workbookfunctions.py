from datetime import datetime  #to get current time
import logging
import os  #for directory stuff
#import pathlib
import sys

#this houses the functions for dealing with the workbooks.
from openpyxl import load_workbook  #for working with excel workbooks
#from pyathena import connect    #for accessing data through athena
#from pyathena.util import as_pandas  #to get panda cursor
from openpyxl.workbook import defined_name  #for working with the defined names
#from openpyxl.worksheet import protection    #for sheet protection  #Argh, haven't gotten this working
from openpyxl.utils import absolute_coordinate
from openpyxl.utils import get_column_letter
import pandas as pd
from itertools import islice

import cpuc.params as params

#import params #parameters file to start the process

class WorkbookDef:
    def __init__(self):
        self.control_dict = {}
        self.range_names = {}
        self.range_names_bad = {}
        self.hidecolors = set()
        self.problems = set()
        self.ok = False
        self.wb = object

class DName:
    def __init__(self):
        self.sheet = ''
        self.range = ''
        self.prefix = ''
        self.name = ''
        self.outtable = ''
        self.intable = ''

def loadtemplate(wb): #returns a WorkbookDef object
    #pass in workbook

    tmpl = WorkbookDef()

    #tmpl.control_dict['studyname'] = 'workbookname'
    #folder = 'P:\\Projects\\CPUC10 (Group D - Custom EM&V)\\4 Deliverables\\21 - Other Contractor Coordination\\ginatemp\\'
    filename = 'Draft_Gross_ExPost_Enhanced.xlsx'
        #load these from the template
    #the workbook folder variable should have the study added to it
    #tmpl.control_dict['workbookfolder'] = 'Z:\\Folders\\Projects\\CPUC10 (Group D - Custom EM&V)\\5 Sites\\ExPost2017\\'
    #tmpl.control_dict['workbookfolder'] = 'P:\\Projects\\CPUC10 (Group D - Custom EM&V)\\4 Deliverables\\21 - Other Contractor Coordination\\ginatemp\\sites\\'
    tmpl.control_dict['rngwbfolder'] = 'workbookfolderrange'
    tmpl.control_dict['rngtreatment'] = 'treatmentname'
    #control_dict['workbookfolder'] = folder + 'sites\\'
    tmpl.control_dict['rngstudy'] = 'StudyName'
    tmpl.control_dict['password'] = 'rpassword'

        #check for problems
        #move this to function above that checks template
            #external links
    if (len(wb._external_links)>0):
        #log error and quit?
        logging.warning('%s has external links. Please fix' % filename)
        # use return to exit function
    #move to load template function
    #pull out new filename
    names = wb.defined_names
    #print (type(names))
    #print (names)
    for dn in names.definedName:
        #print (type(dn))
        #print (dn)
        rng = parse_name({dn.name:dn.attr_text},dn.name)
        if dn.name == tmpl.control_dict['rngstudy']:
            #sheetname = dn.attr_text.split("!")[0]
            #rangelocation = dn.attr_text.split("!")[1]
            #print(sheetname + ", "+ rangelocation)
            logging.info('%s, %s' %(rng.sheet, rng.range))
            #verify that range is a single cell
            #if ':' in rangelocation:
            if ':' in rng.range:
                logging.warning('not single cell')
                tmpl.range_names_bad[dn.name] = 'not single cell '
                #exit giving msg about needing single cell range
            #ws = wb[sheetname]
            #newfilename = ws[rangelocation].value
            tmpl.control_dict['study'] = wb[rng.sheet][rng.range].value
            """
            newfilename =  wb[rng.sheet][rng.range].value
            if '.' in newfilename and '.xlsx' not in newfilename:
                print('uh oh wrong extension or dot in filename')
                tmpl.range_names_bad[dn.name] = 'wrong extension or dot in filename '
                #exit giving message about bad filename
            elif '.xlsx' not in newfilename: #no extension add one
                newfilename += '.xlsx'
            print ('newfilename:' + newfilename)
            #tmpl.control_dict['workbookfilename'] = newfilename
            """
        elif dn.name == tmpl.control_dict['rngwbfolder']:
            folder = wb[rng.sheet][rng.range].value + '\\'
            logging.info('starting folder creation for: %s' % folder)
            tmpl.control_dict['workbookfolder'] = folder
        elif dn.name == tmpl.control_dict['rngtreatment']:
            tmpl.control_dict['treatment'] = wb[rng.sheet][rng.range].value
        elif str(dn.localSheetId) != 'None': #check for sheet level scoped ranges, a no no
            tmpl.range_names_bad[dn.name] = 'sheet level scope'
        elif '#REF!' in dn.attr_text: #bad range
            tmpl.range_names_bad[dn.name] = 'bad refers to: ' + dn.attr_text
        elif '[' in dn.attr_text:   #External link
            #maybe use exteernal links list to add woorkbookname to output. number -1 is index
            tmpl.range_names_bad[dn.name] = 'External link:' + dn.attr_text
        else:
            tmpl.range_names[dn.name] = dn.attr_text
        #print ('name: ' + dn.name + ',  sheet scope:' + str(dn.localSheetId) + ', refers to:' + dn.attr_text)

    for ws in wb.worksheets:
        if len(ws._tables) > 0 :
            #print("Worksheet %s include %d tables:" % (ws.title, len(ws._tables)))
            for tbl in ws._tables:
                tmpl.range_names[tbl.name] = "'" + ws.title + "'!" + tbl.ref

    tmpl.wb = wb
    return tmpl

def check_template_soundness(tmpl):
    #check for soundness
    if 'study' not in tmpl.control_dict or 'treatment' not in tmpl.control_dict:
        tmpl.ok = False
        tmpl.problems.add('missing study or treatment')
        logging.warning('missing study or treatment')
    else:
        tmpl.ok = True

    return tmpl
""" Not used anymore
def loaddata(query): #return dataframe
    #load the data from athena
    #returns a pandas dataframe
    logging.info('loaddata: about to query')
    cursor = connect(aws_access_key_id=params.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=params.AWS_SECRET_ACCESS_KEY,
                    s3_staging_dir='s3://cpuc/',
                    region_name='us-west-2').cursor()

    cursor.execute(query)
    df = as_pandas(cursor)
    logging.info('loaddata: query done')
    return df
"""


def openworkbook(path, values = True, usevba = None): #Returns a workbook or false
    """
    will open workbook and return a wb object
    Values is a boolean setting whether it should open with values or formulas
    usevba is a boolean for the keep_vba parameter
    if it fails returns false
    """
    hasvba = '.xlsm' in path
    if usevba is not None:
        hasvba = usevba
    try:
        wb = load_workbook(path, data_only=values, keep_vba=hasvba)
        return wb
    #except expression as identifier:
    except Exception as e:
        print (f'open workbook exception: {e}')
        return False

def writeworkbook(sampleunit, template):
    if True:
        #sampleid = '212' #
        logging.info('sampleunit:\n%s' % sampleunit)
        sampleid = int(sampleunit['SampleID'].values[0])
        logging.info('sampleid: %s' % sampleid)
        #save as new file
        #create site folder if it doesn't exist
        if 'workbookfolder' not in template.control_dict:
            logging.warning('doh, missing folder in dict')
            return 'no folder cannot write workbook'
        folderpath = template.control_dict['workbookfolder'] + '\\' + template.control_dict['study'] + '\\' + str(sampleid) + '\\' + template.control_dict['treatment']
        try:
            os.makedirs(folderpath, exist_ok = True)
        except:
            #log message that folder couldn't be created
            logging.warning('doh, folder cannot be created')
            #exit

        workbookpath = folderpath + '\\' + template.control_dict['study'] + '_' + template.control_dict['treatment'] + '_' + str(sampleid) + '.xlsx'
        logging.info('workbookpath: %s' % workbookpath)
        #put this in try incase file is unavailable (though unlikely at this point)
        #template.wb.save (workbookpath)  #Maybe don't need this save here
        #go through ranges parsing to decide what to do
        #print(template.range_names)

        for item in template.range_names:

            rngname = parse_name(template.range_names, item)
           #the first section has the instruction codes
            #the last section has the field name

            #maybe first check the length of the instruction section to place it in a group
            #if range[0] == 'fd':
            #print(rngname.prefix)
            if rngname.prefix == 'fdi':
                #add tests for the parts, is sheetname in list of sheets, etc?
                ws = template.wb[rngname.sheet.replace("'", '')]
                try:
                    print('range '+ template.range_names[item] + ' (' + rngname.name + ') should have ' + str(sampleunit[rngname.name].iloc[0]))
                    ws[rngname.range].value = sampleunit[rngname.name].iloc[0]

                except:
                    #print('missing field:' +rngname.name)
                    pass
                #ws[rngname.range].protection = Protection(locked=False)
                #ws[rngname.range].style = Style(protection = Protection(locked=False))
            elif rngname.prefix == 'fdh':
                #fill in table of data
                #this assumes data goes down. May need to adjust if we get data going right
                print('in fhd')
                ws = template.wb[rngname.sheet]
                rng = ws[rngname.range]
                rowindex = 1
                print('sampleunit row count{}'.format(len(sampleunit.index)))
                for _, row in sampleunit.iterrows():
                    print('row type is {}'.format(type(row)))
                    if type(rng) is tuple:
                        for x in rng:
                            for cell in x:
                                print('x is {}'.format(x))
                                print('cell is {}'.format(cell))
                                #cell = x[0]
                                print('cell value is {}'.format(cell.value))
                                try:
                                    print('row[cell.value] is {}'.format(row[cell.value]))
                                    #if cell.value and cell.value in row.index:
                                    print('oringal value is {}'.format(ws.cell(cell.row + rowindex, cell.column).value))
                                    ws.cell(cell.row + rowindex, cell.column).value = row[cell.value]
                                    #ws.cell(cell.row + rowindex, cell.column).protection = Protection(locked=False)
                                except:
                                    print('missing field {}'.format(cell.value))
                    else:
                        print('in not tuple, range is {}'.format(rng))
                        for cell in rng:
                            ws.cell[cell.row + rowindex, cell.column].value = row[cell.value]
                            #ws.cell[cell.row + rowindex, cell.column].protection = Protection(locked=False)
                    rowindex += 1

        #hide sheets
        template.hidecolors = get_values(template.wb, template.range_names, 'ch')
        #print('hide colors follows')
        #print(template.hidecolors)
        for ws in template.wb:
            #print('sheet color: ' + str(ws.sheet_properties.tabColor))
            if ws.sheet_properties.tabColor is not None and ws.sheet_properties.tabColor.rgb in template.hidecolors:
                #print('hide sheet')
                ws.sheet_state = 'hidden'

            #lock sheets, set specific properties
            ws.protection.sheet = True
        #save final file
        template.wb.save (workbookpath)
        return 'success with ' + str(sampleid) + ' ' + str(datetime.now())
    else:
        return 'bad template' + str(datetime.now())

def parse_name(namesdict, name):
    #split a defined named into it's parts
    #TODO Get the table if possible. This may be better somewhere else
    defname = DName()
    ##   !!!check for nonstandard range like dynamic
    defname.sheet = namesdict[name].split('!')[0]
    defname.range = namesdict[name].split('!')[1]
    if '.' in name:
        defname.prefix = name.split('.')[0]
        if len(defname.prefix)>5:
            defname.prefix = ''
            defname.name = name.split('.')[0]
        else:
            defname.name = name.split('.')[1]
    else: #no prefix
        defname.name = name.split('.')[0]
    return defname
def get_name_table(namesdict, definedname):
    pass

def get_values(workbook, rangedict, prefix):
    #get cell values for passed name
    values = set()
    for item in rangedict:
        rngname = parse_name(rangedict, item)
        if rngname.prefix == prefix:
            #print('in prefix match')
            #print (rngname)
            #ws = workbook[rngname.sheet]
            #values.add(ws[rngname.range].value)
            values.add( workbook[rngname.sheet][rngname.range].value)

    #print(values)
    return values

def ret_range_value(wb, sheetname, cellrange):
    return wb[sheetname][cellrange]

def writelog(file, msg):
    pass
    '''
    if os.path.fileexists(file):
        with open(file, "a") as myfile:
            myfile.write(msg + ' ' + str(datetime.now()) + '\r\n' )
    else:
        print (f'log file {file} does not exist')
    '''

    '''
    f= open(file,"w+")
    f.write(msg + "\r\n")
    f.close()
    '''
def writelogsheet(wb, msg):
    #Write message with time to logsheet
    #maybe create msg class to have other information like type of message or whatnot
    pass

def rdl(passedcell):
    return passedcell.parent.cell(passedcell.row, passedcell.column -1)

def rdr(passedcell):
    return passedcell.parent.cell(passedcell.row, passedcell.column +1)

def rda(passedcell):
    return passedcell.parent.cell(passedcell.row -1, passedcell.column)

def rdb(passedcell):
    return passedcell.parent.cell(passedcell.row +1, passedcell.column)

rdlist = {'rdl':rdl, 'rdr':rdr, 'rda':rda, 'rdb':rdb}

def createnames(tmpl):
    #create named ranges for the rd class of names (the range defs)
    print('at the start of createnames')
    deletedcolumns = set()
    for item in tmpl.range_names:
        prefix = item.split(".")[0]
        rangeaddr = tmpl.range_names[item]
        if prefix in rdlist:
            sheet = rangeaddr.split('!')[0].replace("'","")
            ws = tmpl.wb[sheet]
            rng = ws[rangeaddr.split('!')[1]]
            if type(rng) is tuple:
                for x in rng:
                    cell = x[0]
                    neighborcell = rdlist[prefix](cell)
                    assignname(cell, neighborcell)
            else:
                cell = rng
                neighborcell = rdlist[prefix](rng)
                assignname(rng, neighborcell)

            # hiding until the workbook error issue can be sorted out
            usedel = False
            if not usedel:
                hidecol(cell)
            else:
                del cell.parent.parent.defined_names[item]
                if not columndeleted(cell, deletedcolumns):
                    print ('deleteing column for ' + cell.coordinate)
                    deletecol(cell)
                    addtocolumndeleted(cell, deletedcolumns)

    return True

def assignname(sourcecell, targetcell):
    if sourcecell.value is None:
        msg = 'missing range name for ' + sourcecell.parent.title + '!' + sourcecell.coordinate
        print (msg)
        writelogsheet(sourcecell.parent.parent, msg)
        logging.warning('Template has problems: %s' % msg)
        #write to log bad range
    else:
    #check to see if named range is already there
        sourcecell.parent.parent.defined_names.delete(sourcecell.value) #does this crash if name isn't there?
    #if it is either delete or edit
    #if not add it
        newaddress = "'" + targetcell.parent.title + "'" + '!' + absolute_coordinate(targetcell.coordinate)
        #print('cell' + cell.coordinate +  ' new address:' + str(newaddress))
        newname = defined_name.DefinedName(name = str(sourcecell.value), attr_text= newaddress)
        sourcecell.parent.parent.defined_names.append(newname)

def hidecol(rng):
    rng.parent.column_dimensions[get_column_letter(rng.column)].hidden = True

def deletecol(rng):
    ### Not sure why but line below cause a corrupt workbook. Recoverable, but prompts recovery.
    rng.parent.delete_cols(rng.column,1)
    #the line below does not cause a problem
    #rng.parent.delete_cols(7,1)
    pass

def colunmentry(cell):
    return cell.parent.title + str(cell.column)

def columndeleted(cell, deletedset):
    #entry = cell.parent.title + str(cell.column)
    return colunmentry(cell) in deletedset

def addtocolumndeleted(cell, deletedset):
    #entry = cell.parent.title + str(cell.column)
    deletedset.add(colunmentry(cell))

def convertwstodf(sheet, headerrow=1, header = None):
    """
    convert data on worksheet to a dataframe
    Pass in worksheet and row of header (it will skip rows before header and get data directly after)
    Optionally pass in a header list to start at an arbitrary place in the worksheet indicated by the headerrow variable
    If the header has fewer columns than the data on the sheet, columns at the end will be dropped
    uses the first column as the index

    """
    cnt = 1
    for i in range(1,5):
        if sheet.cell(headerrow, 1).value is not None:
            break
        sheet.delete_cols(1,1)        
    # print(f'colindex is {cnt}')
   

    data = sheet.values
    index = 0
    colindex = 1 #doesn't seem to allow for adjusting the data that gets pulled from generator, but leaving it anyway

    while index < headerrow:
        cols = next(data)[colindex-1:]
        index +=1    

    if cols.count(None) > 100: #the sheet has data that's not real so have to process manually.
        #figure out column bount
        realend = len(cols)        
        for item in reversed(cols):
            if item is not None:
                print (f'col item is {item}')
                cols = cols[colindex-1:realend]
                break
            realend = realend - 1
        #something like
        datalist = []
        #blankrowlimit = 5
        for row in data:
            tmprow = row[colindex-1:realend]
            if tmprow.count(None) == len(tmprow): #stops at first blank row
                break
            datalist.append(tmprow)
        #tmprow = next(data)[0:realend]
        data = datalist
    else:
        data = list(data)
    print('length of list in convertwstodf: {}'.format(len(data)))
    idx = [r[0] for r in data]
    
    if header:
        cols = header
        data = (islice(r, 0, len(header)) for r in data)
    else:
        data = (islice(r, 0, None) for r in data)
    if idx.count(None) == len(idx): #in case the first column isn't really the index. Maybe trigger on any none's
        idx = None
    df = pd.DataFrame(data, index=idx, columns=cols)

    return df

def getSampleControlFile(filepath = None, wks = 'Assignments', headerrow = 1, asdf = True, usevalues = True, usevba=False):
    """
    returns a dataframe if asdf is true with the passed sheet list from the administration file
    if asdf is false it returns a worksheet

    """
    if filepath == None:
        filepath = params.SAMPLES_PATH + '\\' + params.SAMPLE_CONTROL_FILE

    sheet = wks   
    wb = openworkbook(filepath, values=usevalues, usevba=usevba)
    if wb:
        ws = wb[sheet]
    else:
        #could't open
        return False

    if asdf: #load sheet into DataFrame        
        #headerrow = 1
        #if (wks == 'Manage Assignments' or wks == 'qcchecks'):
        #    headerrow = 2
        df = convertwstodf(ws, headerrow)
        return df
    else:
        return ws

#def convert_dftoDict
