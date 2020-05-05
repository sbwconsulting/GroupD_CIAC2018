from datetime import datetime
import logging
import math
import os.path
import warnings
#import copy

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles.protection import Protection
from openpyxl.utils import absolute_coordinate
from openpyxl.utils import get_column_letter
from openpyxl.utils import rows_from_range
from openpyxl.workbook import defined_name
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation

from cpuc.sharefileapi import ShareFileSession
from cpuc.sharefileapi import SHAREFILE_OPTIONS
import cpuc.params as params
from cpuc.permissions_workbook import Sample
import cpuc.workbookfunctions as wbfunc
from cpuc.workbookfunctions import convertwstodf

# ?? Does this prevent it from being flexible as the template changes and adds new fields

DATA_PREFIXES = [
    'fdi', 
    'fdh',
    'fdhv' 
]
ELIGIBILITY_SAMPLE_FIELDS = [
    'ProjectID',
    'RvwInstallDate',
    'RvwAppVsInstallDate',
    'RvwPaidIncentive',
    'RvwPermit',
]
MEASURE_DATA_RANGES = [
    'fdh.EligibilityItemHeaders',
    'fdh.MeasuresItemHeaders',
]
PREFILL_COLUMNS = {
    'ELIGIBILITY': [
        'SBW_ProjID',
        'ProjectID',
        'ClaimID',
        'MeasDescription',
    ],
    'MEASURES': [
        'SBW_ProjID',
        'ProjectID',
        'ClaimID',
        'MeasDescription',
        'MeasAppType',
        'EUL_Yrs',
        'RUL_Yrs',
        'CalcGrosskW1stBaseline',
        'CalcGrosskWh1stBaseline',
        'CalcGrossTherm1stBaseline',
        'CalcGrosskW2ndBaseline',
        'CalcGrosskWh2ndBaseline',
        'CalcGrossTherm2ndBaseline',
        'UseCategory',
        'UseSubCategory',
        'TechGroup',
        'TechType',
        'EvalInitialRvwNotes',
    ],
}
CALCULATED_COLUMNS = {
    'CalcGrosskW1stBaseline': 'UnitkW1stBaseline',
    'CalcGrosskWh1stBaseline': 'UnitkWh1stBaseline',
    'CalcGrossTherm1stBaseline': 'UnitTherm1stBaseline',
    'CalcGrosskW2ndBaseline': 'UnitkW2ndBaseline',
    'CalcGrosskWh2ndBaseline': 'UnitkWh2ndBaseline',
    'CalcGrossTherm2ndBaseline': 'UnitTherm2ndBaseline',
}
SHOWN_WORKSHEETS = [
    'SITE INFORMATION',
    'ELIGIBILITY',
    'MEASURES',
    'SUPPLEMENTAL REQUESTS',
    'REFERENCES',
    'PROJECT STATUS',
    'Unmanaged Wkshts',
    'Abbreviations&Glossary',
]


class SampleWorkbook(object):
    """
    Represents a sample workbook. Initialized as follows:

    kwargs:
    template     If True, treats workbook as a template file. Otherwise,
                 it is a regular workbook to be filled and saved.
    input_path   Path to .xlsx file that should be opened. If missing,
                 and template=True, opens the default template file
                 and verifies it. If missing, and template=False, stops
                 and reports an error.
    output_path  Destination path where workbook .xlsx file will be
                 saved. If missing, uses a default path generated from
                 the sample id, study, and treatment.
    data         Sample data source, a pandas DataFrame object that
                 contains measurement data for a single sample.
    sample_id    Sample ID. Controls which data is used for prefill,
                 generating the sample workbook's name, and where the
                 sample workbook is saved.
    study        Name of the study. For example, 'ExPost2017'
    treatment    Name of the treatment. For example, 'Review'
    data_only    Corresponds to the data_only arg in openpyxl's
                 load_workbook function. Defaults to True; pass False
                 to preserve formulas instead of reading their values
                 from the last time Excel saved the workbook.

    In a typical workflow, a SampleWorkbook template is first opened
    and verified. If all goes well, the template is reopened as a regular
    workbook and processed in a loop.

        template = SampleWorkbook(template=True)
        try:
            template.verify()
        except RuntimeError:
            # oops, template has problems

        # retrieve data
        for _, sample_data in sample_groups:
            workbook = SampleWorkbook()
            workbook.prefill(sample_data)
            workbook.protect_and_hide()
            workbook.add_data_validation()
            workbook.save()
    """
    def __init__(self, template=False, **kwargs):
        self.template = template

        self.input_path = kwargs.get('input_path')
        self.output_path = kwargs.get('output_path')
        self.data = kwargs.get('data')

        self.sample_id = kwargs.get('sample_id')
        self.study = kwargs.get('study')
        self.treatment = kwargs.get('treatment')
        self.use_vba = kwargs.get('use_vba')
        try:
            self.data_only = kwargs['data_only']
        except KeyError:
            self.data_only = True

        self.control_dict = {}
        self.range_names = {}
        self.range_names_bad = {}
        self.hidecolors = set()
        self.problems = set()
        self.sample_data = {}
        self.measures = []
        self.tabledefs = []
        self.tablemap = []
        self.fieldlist = []
        self.ok = False
        self.wb = None
        self.loaded = False
        self.sfsession = ShareFileSession(SHAREFILE_OPTIONS)

        self._open()

    def __str__(self):
        if self.template:
            name = 'template'
        else:
            name = self.sample_id
        return '<SampleWorkbook {}>'.format(name)

    @property
    def input_file_path(self):
        if not self.input_path:
            if self.template:
                return params.TEMPLATE_WORKBOOK
            else:
                msg = 'Input path must be specified if not opening a template'
                raise RuntimeError(msg)
        return self.input_path

    @property
    def output_file_name(self):
        return '{}{}_{:0>4}.xlsx'.format(
            self.study,
            self.treatment,
            self.sample_id,
        )

    @property
    def output_file_path(self):
        if self.output_path:
            return self.output_path

        path = os.path.join(
            self.sample.treatment_folder,
            self.output_file_name,
        )
        return self.sample.add_root_to_path(
            params.SAMPLES_PATH,
            path,
        )

    @property
    def sample(self):
        return Sample({
            'SampleID': self.sample_id,
            'Study': self.study,
            'Treatment': self.treatment,
            #'ProjectID' : self.project_id, #Breaks it for some reason. So disable for now.
        })

    def _open(self):
        """
        Open the Excel workbook.

        returns:
        openpyxl Workbook object
        """
        # Suppress unsupported Data Validation warning
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            #logging.warning('INFO - Opening workbook for %s' % self.sample_id)
            #TODO sort out ow to really get this vba property
            hasvba = '.xlsm' in self.input_file_path #this will not work
            # name = self.output_file_name
            # hasvba = '.xlsm' in name #switched because input is opbject
            if self.use_vba is not None:
                hasvba = self.use_vba
            if isinstance(self.input_file_path, str):
                try:                               
                    file_item = self.sfsession.get_io_version(self.input_file_path)                        
                except Exception:
                    logging.critical('Trouble opening workbook %s' % self.sample_id)
                    #TODO figure out how to cancel out of this
                    return False
                else:                
                    ioitem = file_item.io_data
            else:
                ioitem = self.input_file_path
            try:
                self.wb = load_workbook(ioitem, data_only=self.data_only, keep_vba=hasvba)
            except Exception as e:
                logging.critical(f'Trouble opening workbook {self.sample_id}: {e}')
                #TODO figure out how to cancel out of this
                return False
        
        self._load()

    def _load(self):
        """
        Loads control information from the workbook into local data
        structures.
        """
        self.control_dict['rngwbfolder'] = 'workbookfolderrange'
        self.control_dict['rngtreatment'] = 'treatmentname'
        self.control_dict['rngstudy'] = 'StudyName'
        self.control_dict['password'] = 'rpassword'
        self.control_dict['t_version'] = 'templateversion'        

        logging.info('Loading template')
        names = self.wb.defined_names
        for dn in names.definedName:
            rng = DefinedName({dn.name: dn.attr_text}, dn.name)
            if dn.is_external: #it's a worksheet scope not workbook scope?
                continue
            elif dn.name == self.control_dict['rngstudy']:
                logging.info('%s, %s' %(rng.sheet, rng.range))
                # confirm that range is a single cell
                # if ':' in rangelocation:
                if ':' in rng.range:
                    logging.warning('not single cell')
                    self.range_names_bad[dn.name] = 'not single cell'
                    #exit giving msg about needing single cell range
                #ws = wb[sheetname]
                #newfilename = ws[rangelocation].value
                self.control_dict['study'] = self.wb[rng.sheet][rng.range].value
                self.study = self.control_dict['study']
            elif dn.name == self.control_dict['rngwbfolder']:
                try:
                    folder = self.wb[rng.sheet][rng.range].value + '\\'
                # logging.info('starting folder creation for: %s' % folder)
                    self.control_dict['workbookfolder'] = folder
                except:
                    pass
            elif dn.name == self.control_dict['rngtreatment']:
                self.control_dict['treatment'] = self.wb[rng.sheet][rng.range].value
                self.treatment = self.control_dict['treatment']
            elif dn.name == self.control_dict['t_version']:
                self.version = self.wb[rng.sheet][rng.range].value
            elif str(dn.localSheetId) != 'None': #check for sheet level scoped ranges, a no no
                self.range_names_bad[dn.name] = 'sheet level scope'
            elif '#REF!' in dn.attr_text: #bad range
                self.range_names_bad[dn.name] = 'bad refers to: {}'.format(dn.attr_text)
            elif '[' in dn.attr_text:   #External link
                #maybe use exteernal links list to add woorkbookname to output. number -1 is index
                self.range_names_bad[dn.name] = 'External link: {}'.format(dn.attr_text)
            else:
                self.range_names[dn.name] = dn.attr_text

        for ws in self.wb.worksheets:
            if len(ws._tables) > 0 :
                #print("Worksheet %s include %d tables:" % (ws.title, len(ws._tables)))
                for tbl in ws._tables:
                    self.range_names[tbl.name] = '{}!{}'.format(
                        ws.title,
                        tbl.ref
                    )

        self.loaded = True

    def verify(self):
        if self.template:
            self._check_template_soundness()
            self._check_data_params()

        if self.ok:
            logging.info('Workbook verified, no issues found.')
        else:
            msg = 'Workbook verified, problems found. See log for details'
            logging.info(msg)
            raise RuntimeError(msg)

    def _check_template_soundness(self):
        logging.info('Checking template soundness')
        ok = True

        if 'study' not in self.control_dict:
            logging.warning('Template is missing study')
            self.problems.add('Missing study')
            ok = False

        if 'treatment' not in self.control_dict:
            logging.warning('Template is missing treatment')
            self.problems.add('Missing treatment')
            ok = False

        if len(self.wb._external_links) > 0:
            #Add check to see if it's from setup reqs or not
            #TODO switch to sf item name
            logging.warning(
                '%s has external links. Please fix' % self.input_file_path
            )
            self.problems.add('Has external links')
            ok = False
        #build field list and check for field's with None
        
        self.get_table_defs()
        self.generate_table_map(writetosheet=False)                
        self.generate_field_list()
        
        missingtablenames = [x['Source'] for x in self.tabledefs if x['TableName'] is None]
        if not missingtablenames == []:
                msg = f'Missing table name{missingtablenames.join(", ")}'
                logging.warning(msg)
                self.problems.add(msg)
                ok = False
        for item in self.fieldlist:
            if item['field'] == None or '=' in item['field']:
                msg = f'Missing field name: {item["field"]}'
                logging.warning(msg)
                self.problems.add(msg)
                ok = False
        
        self.ok = ok

        if ok:
            logging.info('Template is sound')
        else:
            logging.warning('Template has problems: %s' % self.problems)
            print('Template has problems: %s' % self.problems)

    def _check_data_params(self):
        """
        Check that all field parameters needed for processing exist.
        """
        # not sure if this will be specific to study or if can generalize it
        ok = True
        if not params.STUDY_TABLE:
            ok = False
        if not params.SAMPLE_ID_FIELD:
            ok = False
        if not params.QUERY_SAMPLE_IDS:
            ok = False

        if not ok:
            self.ok = False
            msg = 'Not all query data specified in params file.'
            logging.warning(msg)

        self.ok = ok

    def _get_common_field(self, name):
        """
        Retrieves a value common to all measurements in the sample.
        Really just grabs the value of the named field from the first
        row in the sample, so it's up to the caller not to use this on
        fields that contain different data for different measurements.

        args:
        name - name of the field to retrieve
        """
        return self.data[name].iloc[0]

    def get_sampleid(self, sampleid):
        if is_number(sampleid):
            return int(sampleid)
        else:
            return sampleid

    def prefill(self, data):
        """
        Fill in autofill fields from supplied data.

        args:
        data  Pandas DataFrame containing sample data
        """
        self.data = data
        sid = self._get_common_field('SampleID')
        if is_number(sid):
            self.sample_id = int(sid)
        else:
            self.sample_id = sid
        #self.project_id = self._get_common_field('ProjectID')
        logging.info('Prefilling sample id %s' % self.sample_id)
        #check for output path
        outfolder = self.sample.add_root_to_path(
            params.SAMPLES_PATH,
            self.sample.treatment_folder)
        if not os.path.exists(outfolder):
            print (f'no folder for {outfolder}')
            logging.critical(f'Missing folder for {outfolder}, no workbook generated')
            return False

        result = self.writeworkbook(self.data)
        logging.info('%s' % str(result))
        return True

    def collect_data(self):
        """
        Collect workbook data into a dict of lists of dictionaries in self.data.

        Format
                {'table1':
                    [
                        {'SampleID': 'SBW448', 'Measurement1': 87.18, ...},
                        {'SampleID': 'SBW392', 'Measurement1': 28.19, ...},
                        ...
                    ]
                }
                {'table2':
                    [
                        {'SampleID': 'SBW448', 'Measurement1': 87.18, ...},
                        {'SampleID': 'SBW392', 'Measurement1': 28.19, ...},
                        ...
                    ]
                }

        """

        
        data = {}
        for item in self.range_names:
            rngname = DefinedName(self.range_names, item)
            try:
                ws = self.wb[rngname.sheet.replace("'", '')]
            except KeyError:
                #probably a dynamic range, but in any event, sheet doesn't exist so skip
                continue
            if item == 'fdi.SampleID':
                print("debug")
            if rngname.prefix in DATA_PREFIXES:
                tablename = self.get_range_table(rngname.prefix, rngname.name, 'output_tbl')
                #temp
                #if tablename == 'Tracker':
                #    print('tracker')
                tablelist = data.get(tablename)
                if tablename is not None:                    
                    try:
                        tableprops = next(x for x in self.tabledefs if x["TableName"] == tablename)
                        tableindex = tableprops['IndexField']
                    except:
                        print(f'FATAL Error: Missing table {tablename} for {item}')
                        tableindex = None    
                else:
                    #If there is no tablename, then skip loading the data
                    #TODO make sure this doesn't break everything
                    continue
                    tableindex = None
                #    print(f'no output table for {item}')
            if rngname.prefix == 'fdi':
                #column_name = rngname.name
                if tablelist is None:
                    tablelist = []
                    #TODO figure out how to look up index in these cases in case it's not sampleID level 
                    if tableindex != 'SampleID':
                        if tableindex is not None:
                            print(f'uh oh table index {tableindex} is not sample')
                    index = dict({tableindex: self.sample_id})
                    #index[tableindex] = self.sample_id
                    tablelist.append(index)
                for table in tablelist:
                    table[rngname.name] = ws[rngname.range].value
                '''
                # Merge with data already collected for this index
                if tablelist is not None:
                    existing_data = list(
                        item for item in tablelist
                        if item[tableindex] == row_data[tableindex]
                    )
                if existing_data:
                    existing_data[0].update(row_data)
                else:
                '''
                #tablelist.append(row_data)

                data[tablename] = tablelist

            elif rngname.prefix == 'fdh' or rngname.prefix == 'fdhv':
                if rngname.prefix == 'fdh':
                    offsettype = 'row'
                else:
                    offsettype = 'col'
 
                rng = rngname.range

                header_cells = ws[rng][0]
                header_names = self.get_values_in_cells(header_cells)
                '''
                #workaround for ranges that don't have an index in them
                #doesn't work because the processing code needs a contiguous range
                if 'DetailID' not in header_names:
                    #for the specific ciac 2018 case
                    if self.study == 'CIAC2018':
                        header_cells = header_cells + (ws['$D$8'],)
                        header_names = self.get_values_in_cells(header_cells)
                '''    

                offset = 2
                originaloffset = offset
                blankcount = 0
                while True:
                    row_data = {}
                    cells = self.get_cells_for_offset(offset, offsettype, header_cells)
                    if not cells[0].value:
                        blankcount += 1
                        if blankcount > 10:
                            break
                    for index, cell in enumerate(cells):
                        header = header_names[index]                        
                        row_data[header] = cell.value
                    # Merge with data already collected for this index
                    existing_data = None
                    measureindex = None
                    
                    if tablelist is None:
                        tablelist = []
                    else:
                        #this is a work around until index is looked up
                        if not tableindex in row_data:                            
                            #special case for ciac                            
                            if 'measures' in cells[0].parent.title.lower():
                                measureindex = 'DetailID'
                                tableindex = measureindex
                                idcell = self.get_cells_for_offset(offset, offsettype, [ws['D8']])
                            else:
                                tableindex = 'SampleID'
                            #for cases where there isn't an id in the dataset
                            if not tableindex in row_data:
                                if tableindex == 'SampleID':
                                    row_data['SampleID'] = self.sample_id
                                if tableindex == measureindex:
                                    row_data[measureindex] = idcell[0].value
                        
                        #might be able to change this to a boolean sonce not using the list anymore
                        existing_data = list(
                            item for item in tablelist
                            if item[tableindex] == row_data[tableindex]
                        )
                    if tableindex not in row_data:
                        #missing an index field, add one
                        #special case for ciac                        
                        if 'measures' in cells[0].parent.title.lower():
                            measureindex = 'DetailID'
                            tableindex = measureindex
                            idcell = self.get_cells_for_offset(offset, offsettype, [ws['D8']])
                        else:
                            tableindex = 'SampleID'
                        #for cases where there isn't an id in the dataset
                        if not tableindex in row_data:
                            if tableindex == 'SampleID':
                                row_data['SampleID'] = self.sample_id
                            if tableindex == measureindex:
                                row_data[measureindex] = idcell[0].value
                        #missing an index field, skip to next
                        #break
                    
                    if row_data[tableindex] != None: #skip rows with blank index
                        if existing_data:
                            try:
                                tablelist[offset - originaloffset].update(row_data)
                            except:
                                msg = f'SampleID {self.sample_id} more data than tablelist for {offset}: {row_data}'
                                print(msg)
                                logging.warning(msg)
                        else:
                            tablelist.append(row_data)
                    #else:
                    #    print(f'skipping row {row_data}')
                    offset += 1
                
                data[tablename] = tablelist

        self.data = data

    def get_range_table(self, prefix, name, tabletype):
        '''
        returns the name of the table matching the passed parameters.
        Tabletype options are output_tbl, input_tbl
        Returns none if no match is found
        '''
        rangeinfo = next((item for item in self.tablemap if item['prefix']==prefix and item['range']==name),None)
        if rangeinfo is None:
            return None
                
        rangeinfo = rangeinfo.get(tabletype)

        return rangeinfo

    def get_table_data(self, table):
        """
        Pulls out the list from the self.data dictionary. 
        """
        
        
       
        return self.data.get(table)


    def collect_data_D0(self):
        """
        Collect workbook data into python data structures. Grabs measure
        data first, because sample-level data pulls from the first row
        of the ELIGIBILITY worksheet.
        """
        self.collect_measure_data()
        self.collect_sample_data()

    def collect_sample_data(self):
        """
        Fills dict of fields that contain sample-level data. Relies on
        existence of self.measure_data (from collect_measure_data()) for
        collection of sample-level data from the ELIGIBILITY worksheet.
        """
        logging.info("Collecting sample-level data")
        if self.sample_id:
            self.sample_data['SampleID'] = self.sample_id
        for item in self.range_names:
            rngname = DefinedName(self.range_names, item)
            ws = self.wb[rngname.sheet.replace("'", '')]
            if rngname.prefix == 'fdi':
                try:
                    column_name = rngname.name
                    self.sample_data[column_name] = ws[rngname.range].value
                    '''
                    logging.info('Found %s with value %s' % (
                        column_name,
                        self.sample_data[column_name],
                    ))
                    '''
                except KeyError:
                    logging.info('missing field: %s' % rngname.name)
        try:
            first_measure = self.measure_data[0]
        except AttributeError as e:
            msg = (
                'Measure data has not been collected. Run'
                ' collect_measure_data() before collect_sample_data().'
            )
            logging.error(msg)
            raise(e)
        for field in ELIGIBILITY_SAMPLE_FIELDS:
            self.sample_data[field] = first_measure[field]

        logging.info("Done collecting sample-level data")

    def collect_measure_data(self):
        """
        Fills list of dicts containing measurement data.
        """
        logging.info("Collecting measure-level data")

        data = []
        for range_name in MEASURE_DATA_RANGES:
            dn = DefinedName(self.range_names, range_name)
            ws = self.wb[dn.sheet.replace("'", '')]
            rng = dn.range

            header_cells = ws[rng][0]
            header_names = self.get_values_in_cells(header_cells)
            offset = 1
            while True:
                row_data = {}
                cells = self.get_cells_for_row_offset(offset, header_cells)
                if not cells[0].value:
                    break
                for index, cell in enumerate(cells):
                    header = header_names[index]
                    '''
                    logging.info('Found %s with value %s' % (
                        header,
                        cell.value
                    ))
                    '''
                    row_data[header] = cell.value
                # Merge with data already collected for this ClaimID
                existing_measurements = list(
                    measure for measure in data
                    if measure['ClaimID'] == row_data['ClaimID']
                )
                if existing_measurements:
                    existing_measurements[0].update(row_data)
                else:
                    data.append(row_data)
                offset += 1

        self.measure_data = data
        logging.info("Done collecting measure-level data")

    def get_cells_for_row_offset(self, row_offset, cells):
        """
        Given a list of cells, returns the cells in the offset row
        beneath that list, bounded by the left and right columns of
        the cells in the list. Only works with single-row, contigous
        ranges.

        args:
        row_offset  Offset from the input list of cells. The first row
                    row beneath the cells is row_offset=1.

        cells       List of cells from which offset and column bounds
                    are calculated.

        returns:
        List of cells in the new row.
        """
        if cells[0].row != cells[-1].row:
            msg = 'Range must contain only one row'
            raise ValueError(msg)
        for index, cell in enumerate(cells):
            try:
                if cell.column + 1 != cells[index + 1].column:
                    msg = 'Range must be contiguous'
                    raise ValueError(msg)
            except IndexError:
                pass  # last cell in range
        return [cell.offset(row=row_offset) for cell in cells]

    def get_cells_for_offset(self, offset, offsettype, cells):
        """
        Given a list of cells, returns the cells in the offset row/column
        beneath/right of that list, bounded by the columns/rows of
        the cells in the list. Only works with single-row/column, contigous
        ranges.

        args:
        offset  Offset from the input list of cells. The first row
                    row beneath the cells is row_offset=1.
        
        offsettype  row or col

        cells       List of cells from which offset and column bounds
                    are calculated.

        returns:
        List of cells in the new offset.
        """
        if offsettype == 'row':
            coloffset = 0
            rowoffset = offset
            if cells[0].row != cells[-1].row:
                msg = 'Range must contain only one row'
                raise ValueError(msg)
        elif offsettype == 'col':
            coloffset = offset
            rowoffset = 0
            if cells[0].column != cells[-1].column:
                msg = 'Range must contain only one column'
                raise ValueError(msg)
        
        for index, cell in enumerate(cells):
            try:
                if ((offsettype == 'row' and cell.column + 1 != cells[index + 1].column)
                    or (offsettype == 'col' and cell.row + 1 != cells[index + 1].row)):
                    msg = 'Range must be contiguous'
                    raise ValueError(msg)
            except IndexError:
                pass  # last cell in range
        return [cell.offset(row=rowoffset, column=coloffset) for cell in cells]

    def get_values_in_cells(self, cells):
        """
        Returns a list of values within a given list of cells.
        """
        return [cell.value for cell in cells]

    def get_values_in_range(self, rangename):
        """
        Returns a list of values within the range
        """
        
        dn = DefinedName(self.range_names, rangename)
        ws = self.wb[dn.sheet.replace("'", '')]
        rng = dn.range

        range_cells = ws[rng][0]
        range_data = self.get_values_in_cells(range_cells)
        return range_data

    def get_cells_in_range(self, rangename):
        """
        Returns the cells within the range
        """
        
        dn = DefinedName(self.range_names, rangename)
        ws = self.wb[dn.sheet.replace("'", '')]
        rng = dn.range

        range_cells = ws[rng][0]
        return range_cells

    def writeworkbook(self, sampleunit):
        """
        Write data into workbook according to ranges contained in
        Defined Names in the workbook.

        args:
        sampleunit  Pandas DataFrame containing just the data for
                    one sample
        """
        logging.debug('sampleunit:\n%s' % sampleunit)
        sampleid = int(sampleunit['SampleID'].values[0])
        logging.info('sampleid: %s' % sampleid)

        #get fieldlist if it doesn't exist
        if self.fieldlist == []:
            #self.build_field_list(,,'input')
            self.generate_field_list()
        if self.tablemap == []:
            self.generate_table_map(writetosheet=False)

        #limit fieldlist to inputs
        #inputfieldlist = [d for d in self.fieldlist if d['type'] == 'input']
        
        #go through ranges parsing to decide what to do
        logging.info('self.range_names: %s' % self.range_names)
        for item in self.range_names:
            #logging.info('item: %s' % item)

            rngname = DefinedName(self.range_names, item)
            # Some worksheet names are surrounded by single quotes,
            # which cause errors later when attempting to use
            # them as keys for looking up a worksheet in the
            # workbook. For example: ws.["'FOO'"] is not the same as
            # ws["FOO"].
            try:
                ws = self.wb[rngname.sheet.replace("'", '')]
            except KeyError:
                #probably a dynamic range, but in any event, sheet doesn't exist so skip
                continue

            #the first section has the instruction codes
            #the last section has the field name
            if rngname.prefix == 'fdi':
                #add tests for the parts, is sheetname in list of sheets, etc?

                try:
                    column_name = rngname.name
                    # If ProjectID column is empty, use ApplicationCode
                    '''
                    if column_name == 'ProjectID':
                        try:
                            if math.isnan(sampleunit[column_name]):
                                column_name = 'ApplicationCode'
                        except TypeError:
                            # Tried to do math.isnan on string, so ProjectID
                            # must not be empty. Leave it alone.
                            pass
                    '''
                    ws[rngname.range].value = sampleunit[column_name].iloc[0]
                    """
                    logging.info('range %s (%s) should have %s' % (
                        self.range_names[item],
                        rngname.name,
                        str(sampleunit[column_name].iloc[0]),
                    ))
                    """

                except KeyError:
                    logging.info('missing field: %s' % rngname.name)

                #ws[rngname.range].protection = Protection(locked=False)
                #ws[rngname.range].style = Style(protection = Protection(locked=False))
            elif rngname.prefix == 'fdh' or rngname.prefix == 'fdhv':
                if rngname.prefix == 'fdh':
                    orientation = 'row'
                elif rngname.prefix == 'fdhv':
                    orientation = 'col'
            
                rngfieldlist = [x['field'] for x in self.fieldlist if x['range'] == rngname.name and x['type'] == 'input']
                if rngfieldlist == []: #skip output only ranges
                    continue
                #fill in table of data                
                logging.info('in fhd')
                rng = ws[rngname.range]
                startindex = 2 #to skip the data type row
                rowindex = startindex
                colindex = startindex
                logging.info('sampleunit row count %s' % len(sampleunit.index))
                for _, row in sampleunit.iterrows():
                    if orientation == 'row':
                        colindex = 0
                    elif orientation == 'col':
                        rowindex = 0

                    if type(rng) is tuple:
                        for x in rng:
                            for cell in x:
                                logging.debug('x is %s' % (x,))
                                logging.debug('cell is %s' % cell)
                                logging.debug('cell value is %s' % cell.value)
                                try:
                                    column_name = cell.value
                                    # Skip columns that we don't need to fill.
                                    if column_name not in rngfieldlist:
                                        continue
                                    #TODO find another way to skip unused  columns?
                                    #use the sampleunit cols as the restrictor?
                                    #if column_name not in PREFILL_COLUMNS[rngname.sheet]:
                                    #    continue
                                    # If ProjectID column is empty, use ApplicationCode
                                    if column_name == 'ProjectID':
                                        try:
                                            if math.isnan(row[cell.value]):
                                                column_name = 'ApplicationCode'
                                        except TypeError:
                                            # Tried to do math.isnan on string, so ProjectID
                                            # must not be empty. Leave it alone.
                                            pass
                                    #if column_name in CALCULATED_COLUMNS:
                                    #    column_value = self.calculate_gross_savings_D0(
                                    #        row,
                                    #        column_name,
                                    #    )
                                    #else:
                                    column_value = row[column_name]
                                    ws.cell(cell.row + rowindex, cell.column + colindex).value = column_value
                                    logging.debug('column_name is %s' % (
                                        column_name
                                    ))
                                    logging.debug('original value is %s' % (
                                        ws.cell(
                                            cell.row + rowindex,
                                            cell.column + colindex
                                        ).value
                                    ))
                                    # XXX The following line should lock
                                    # the cell, but openpyxl doesn't
                                    # write this value to the saved
                                    # workbook. Currently relying on
                                    # the template to have these cells
                                    # locked already.
                                    # ws.cell(
                                    #     cell.row + rowindex,
                                    #     cell.column,
                                    # ).protection = Protection(locked=False)
                                except KeyError:
                                    logging.info('missing field %s' % cell.value)
                    else:
                        logging.info('in not tuple, range is %s' % rng)
                        for cell in rng:
                            ws.cell[cell.row + rowindex, cell.column + colindex].value = row[cell.value]
                            # XXX Again, this should lock the cell, but doesn't.
                            # ws.cell[cell.row + rowindex, cell.column].protection = Protection(locked=True)
                    rowindex += 1
                    colindex += 1
        return True

    def get_values(self, rangedict, prefix):
        #get cell values for passed prefix
        values = set()
        for item in rangedict:
            rngname = DefinedName(rangedict, item)
            if rngname.prefix == prefix:
                #print('in prefix match')
                #print (rngname)
                #ws = workbook[rngname.sheet]
                #values.add(ws[rngname.range].value)
                values.add(self.wb[rngname.sheet][rngname.range].value)

        #print(values)
        return values

    def calculate_gross_savings_D0(self, row, gross_column):
        unit_column = CALCULATED_COLUMNS[gross_column]
        num = row[unit_column]
        num_units = row['NumUnits']
        return num_units * num

    def add_data_validation_D0(self):
        """
        Hard coded for D0 data validation lists only.
        Hack to work around openpyxl's refusal to correctly read in some
        data validations when opening the workbook. List validations
        pointing to pivot tables seem to work ('EUL Lists' worksheet),
        but list validations pointing directly at cell ranges (as in
        'Lists') are discarded when opening a workbook.
        """
        for ws in self.wb.worksheets:
            if ws.title == 'ELIGIBILITY':
                self.create_data_validation(
                    ws,
                    formula1='=Lists!$B$4:$B$7',
                    ranges=[
                        'F19:I19',
                        'J19:O40',
                    ],
                )
            if ws.title == 'MEASURES':
                self.create_data_validation(
                    ws,
                    formula1='=Lists!$N$4:$N$9',
                    ranges=['AP9:AP40'],
                )
                for x in 'QRSTUVWX':
                    self.create_data_validation(
                        ws,
                        formula1='=Lists!$B$4:$B$5',
                        ranges=['A{0}9:A{0}40'.format(x)],
                    )
                for x in 'BDFHJL':
                    self.create_data_validation(
                        ws,
                        formula1='=Lists!$AF$4:$AF$13',
                        ranges=['B{0}9:B{0}40'.format(x)],
                    )
            if ws.title == 'PROJECT STATUS':
                self.create_data_validation(
                    ws,
                    formula1='=Lists!$AD$4:$AD$7',
                    ranges=['E11:E22'],
                )

    def create_data_validation(self, ws, formula1, ranges):
        """
        Creates a list DataValidation that points to a specified formula
        range, then applies it to specified ranges in a worksheet.

        args:
        ws        Worksheet to which the validation should be attached.
        formula1  Formula pointing to the range of values to use in
                  the list. For example, '=Lists$B$4:$B$7'
        ranges    Ranges in the current worksheet to which the
                  validation should apply.

        """
        dv = DataValidation(
            type='list',
            formula1=formula1,
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        for rng in ranges:
            dv.add(rng)

    def protect_and_hide(self):
        """
        Protects sheets from inadvertent editing, and hides sheets that
        should be hidden.
        """
        # Retrieve tab colors to be hidden.
        self.hidecolors = self.get_values(self.range_names, 'ch')

        for ws in self.wb:
            if ws.sheet_properties.tabColor:
                if ws.sheet_properties.tabColor.rgb in self.hidecolors:
                    ws.sheet_state = 'hidden'

            ws.protection.sheet = True
            #the lines below allow for the resizing of the rows and columns
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False

            # XXX The following line should set a sheet password,
            #     but openpyxl doesn't write it to the workbook
            #     when it saves. The worksheets are still locked,
            #     but there isn't a password to unlock them.
            # ws.protection.password = 'Castor'

    def get_table_defs(self):
        """
        Load the table defs from the tabledefs sheet
        """

        #TODO add qc checks
        # what errors should be checked for?
        rngtabledef = 'a.TableDefs'
        if rngtabledef in self.range_names:
            defrng = DefinedName(self.range_names, rngtabledef)
            ws = self.wb[defrng.sheet]
            rng = ws[defrng.range]
            tbllist = []
            tbllist = convert_table_to_list(rng)
            self.tabledefs = tbllist            

    def generate_table_map(self, writetosheet = True):
        """
        writes all the data ranges and their tables to a TableMap sheet
        """
        deftblrangename = 'a.DefaultDataTableOut'
        mapheaderrangename = 'a.TableMapHeader'
        hidecolorrangename = 'ch.hidepurple'

        if deftblrangename in self.range_names:
            defrng = DefinedName(self.range_names, deftblrangename)
            #rng = self.wb[defrng.sheet][defrng.range]
            defaultInputTable = self.wb[defrng.sheet][defrng.range].offset(column = 1).value
            defaultOutputTable = self.wb[defrng.sheet][defrng.range].value
        else:
            print('missing def tbl range')
            return False

        if hidecolorrangename in self.range_names:
            hiderng = DefinedName(self.range_names, hidecolorrangename)
            #hidecolor = '#' + self.wb[hiderng.sheet][hiderng.range].value
            hidecolor = self.wb[hiderng.sheet][hiderng.range].value.replace('#','')
        else:
            #hidecolor = '#FF7030A0'
            hidecolor = 'FF7030A0'


        #Get list of defs
        if mapheaderrangename in self.range_names:            
            #maphdrrng = DefinedName(self.range_names, mapheaderrangename)
            #rng = maphdrrng.range
            #header_cells = selfws[rng][0]
            header_cells = self.get_cells_in_range(mapheaderrangename)
            #header_cells = self.get_values_in_range(mapheaderrangename)
            #get table map def
            tableranges = []
            map_data = []
            offset = 1
            #
            while True:
                row_data = {}
                cells = self.get_cells_for_row_offset(offset, header_cells)
                if not cells[0].value:
                    break
                for index, cell in enumerate(cells):
                    header = header_cells[index].value                    
                    row_data[header] = cell.value
                    if index == 0:
                        rangeheadername = header
                        tableranges.append(cell.value)
                    elif index ==1:
                        outhdrname = header
                    elif index == 2:
                        inhdrname = header
                
                map_data.append(row_data)
                offset += 1
        else:
            print('table def missing')
            return False

        if writetosheet:
            #drop sheet if it's there
            mapsheet = 'TableMap'
            if mapsheet in self.wb.sheetnames:
                self.wb.remove(self.wb[mapsheet])

            self.wb.create_sheet(mapsheet)
            ws = self.wb[mapsheet]
            ws.sheet_properties.tabColor = hidecolor     #set sheet color to the hide color
            inputcol = 3
            outputcol = 2
            #if writetosheet:
            ws.cell(1,1).value  = 'Range Name'
            ws.cell(1,inputcol).value  = 'Input Table'
            ws.cell(1,outputcol).value  = 'Output Table'                

        offset = 1
        rows_list = []        
        for rng in self.range_names:
            rname = DefinedName(self.range_names, rng)
            if rname.prefix in DATA_PREFIXES:
                #write it to sheet and list
                tbldict = {}
                if writetosheet:
                    ws.cell(1+offset,1).value = rng
                #lookup up input
                if rng in tableranges:                 
                    for item in map_data:
                        if item[rangeheadername] == rng:
                            inputtbl = item[inhdrname]
                            outputtbl = item[outhdrname]
                            if writetosheet:
                                ws.cell(1+offset,inputcol).value = inputtbl
                                ws.cell(1+offset,outputcol).value = outputtbl
                        continue
                else:
                    inputtbl = defaultInputTable
                    outputtbl = defaultOutputTable
                    if writetosheet:
                        ws.cell(1+offset,inputcol).value = inputtbl
                        ws.cell(1+offset,outputcol).value = outputtbl

                #add to list
                tbldict.update({
                    'range': rname.name,
                    'prefix': rname.prefix,
                    'input_tbl':  inputtbl,
                    'output_tbl': outputtbl
                })
                rows_list.append(tbldict)
                offset += 1
        self.tablemap = rows_list

    def build_field_list(self, tablename, tabletype):
        """
        Put together a complete list of fields for the passed table and table type
        tabletype is either input or output

        """
        #print(f'start of build field list. length is {len(self.fieldlist)}, adding {tablename}')
        if self.tablemap is None or self.tablemap == []:
            self.generate_table_map(writetosheet=False)

        #go through Tablemap looking for table
        col = tabletype + '_tbl'
        
        ranges = [{'range':row['range'],'prefix':row['prefix']}  for row in self.tablemap if row[col] == tablename]
        if ranges == []:
            return []
        fieldlist = []
        for cell in ranges:
            fielddict = {}
            if cell['prefix'] == 'fdi':
                fieldtype = self.get_fdi_field_type(fieldname = cell['range'] )
                fielddict.update({
                    'table': tablename,
                    'range': cell['range'],
                    'type': tabletype,
                    'field': cell['range'],
                    'fieldtype':fieldtype,
                })
                fieldlist.append(fielddict)
            elif cell['prefix'] == 'fdh' or cell['prefix'] == 'fdhv':
                #grab range
                if cell['prefix'] ==  'fdh':
                    rowoffset = 1
                    coloffset = 0
                elif cell['prefix'] == 'fdhv':
                    rowoffset = 0
                    coloffset = 1

                rng = DefinedName(self.range_names, cell['prefix'] + '.' + cell['range'])
                headers = self.wb[rng.sheet][rng.range]
                for cell in headers[0]:
                    fielddict = {}
                    fielddict.update({
                        'table': tablename,
                        'range': rng.name,
                        'type': tabletype,
                        'field': cell.value,
                        'fieldtype':  self.wb[rng.sheet].cell(cell.row + rowoffset, cell.column + coloffset).value 
                    })
                    fieldlist.append(fielddict)
                #fields = {{'table': tablename,'field': x} for x in headers}

                #print('process header to get fields')
        if fieldlist != []:            
            #print(f'field list length is {len(self.fieldlist)} before extend')
            self.fieldlist.extend(fieldlist)
            #print(f'field list length is {len(self.fieldlist)} after extend')
        #else:
            #print('no additions to field list')
        #print(f'end of build field list. length is {len(self.fieldlist)}')

    def get_fdi_field_type(self, fieldname):
        """
        lookup the field type for an fdi cell
        returns found type or Text if not found
        """
        try:
            typerange = self.range_names['fdit.' + fieldname]
            rparts = typerange.split('!')
            fieldtype = self.wb[rparts[0].replace("'", "")][rparts[1]].value
        except KeyError as e:
            #print(f'fdi key lookup error for {e}')
            fieldtype = 'Text'
        except Exception as e:
            print(f'fdi lookup error for {fieldname} is {e}')
            fieldtype = 'Text'

        return fieldtype

    def generate_field_list(self):
        """
        Populates self.fieldlist with all the defined fields and their type (inpur or output) and source
        """
        if self.tabledefs == []:
            self.get_table_defs()
         #for each table get the fields        
        for row in self.tabledefs:
            self.build_field_list(row['TableName'], 'input')
            self.build_field_list(row['TableName'], 'output')

    def create_names(self):
        #create named ranges for the rd class of names (the range defs)
        print('at the start of createnames')
        rdlist = {'rdl':rdl, 'rdr':rdr, 'rda':rda, 'rdb':rdb}
        deletedcolumns = set()
        for item in self.range_names:
            prefix = item.split(".")[0]
            rangeaddr = self.range_names[item]
            if prefix in rdlist:
                sheet = rangeaddr.split('!')[0].replace("'","")
                ws = self.wb[sheet]
                rng = ws[rangeaddr.split('!')[1]]
                if type(rng) is tuple:
                    for x in rng:
                        cell = x[0]
                        neighborcell = rdlist[prefix](cell)
                        wbfunc.assignname(cell, neighborcell)
                else:
                    cell = rng
                    neighborcell = rdlist[prefix](rng)
                    wbfunc.assignname(rng, neighborcell)

                # hiding until the workbook error issue can be sorted out
                usedel = False
                if not usedel:
                    wbfunc.hidecol(cell)
                else:
                    del cell.parent.parent.defined_names[item]
                    if not wbfunc.columndeleted(cell, deletedcolumns):
                        print ('deleteing column for ' + cell.coordinate)
                        wbfunc.deletecol(cell)
                        wbfunc.addtocolumndeleted(cell, deletedcolumns)

        self._load() #reload variables so the new names will be there
        return True

    def save(self):
        if self.template and self.input_file_path == self.output_file_path:
            msg = (
                'Cannot overwrite template file. Reopen with'
                ' template=False or pass an output_file_path and try again.'
            )
            raise RuntimeError(msg)
        logging.info('Saving %s' % self.output_file_path)
        
        try:
            self.wb.save(self.output_file_path)
        except PermissionError:
            msg = (
                'PermissionError writing to %s.'
                ' Does somebody have the file open?'
            )
            logging.warning(msg % self.output_file_path)
        #except FileNotFoundError:
        #TODO Currently crashes if path isn't present. 
        # Add in mechanism to create the folder structure
        # or is it better to figure out why the folder isn't there
        # maybe add a check at the beginning of the process

class DefinedName(object):
    def __init__(self, names_dict, name):
        self.sheet = ''
        self.range = ''
        self.prefix = ''
        self.prefixextra = ''
        self.name = ''
        self._parse(names_dict, name)

    def _parse(self, names_dict, name):
        """
        Splits a defined named into its parts

        args:
        names_dict - Dictionary of names in the workbook
        name       - Name of the something
        """
        #split a defined named into its parts
        ##   !!!check for nonstandard range like dynamic because they don't work :(
        if '[' in names_dict[name]:
            #it's a table range, ignore
            return self
            
        self.sheet = names_dict[name].split('!')[0].strip("'")
        try:
            self.range = names_dict[name].split('!')[1]
        except:
            print(f'cannot handle range {name} with refers to of {names_dict[name]}')
            self.range = None
        if '.' in name:
            self.prefix = name.split('.')[0]
            if '_' in self.prefix:
                self.prefixextra = self.prefix.split('_')[1]
                self.prefix = self.prefix.split('_')[0]            
            if len(self.prefix) > 5:
                self.prefix = ''
                self.name = name.split('.')[0]
            else:
                self.name = name.split('.')[1]
        else: #no prefix
            self.name = name.split('.')[0]
        return self

def rdl(passedcell):
    return passedcell.parent.cell(passedcell.row, passedcell.column -1)

def rdr(passedcell):
    return passedcell.parent.cell(passedcell.row, passedcell.column +1)

def rda(passedcell):
    return passedcell.parent.cell(passedcell.row -1, passedcell.column)

def rdb(passedcell):
    return passedcell.parent.cell(passedcell.row +1, passedcell.column)

def convert_table_to_list(therange):
    """ 
    Returns passed in range as a list
    Assumes first row is the header
    """
    rng = therange
    tbllist = []
    for row in rng:           
        #print(f'in iterrows with {row}')
        defdict = {}
        if row[0].row != rng[0][0].row:
            for cell in row:                    
                #print(f'in itercols with {cell}')                                        
                defdict.update({rng[0][cell.column-1].value : cell.value})
            tbllist.append(defdict)
    return tbllist
        
def is_number(s):
    if s is None:
        return False
    if 'Time' in type(s).__name__:        
        return True
    if 'Na' in type(s).__name__:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False